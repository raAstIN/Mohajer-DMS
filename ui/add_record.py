import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import jdatetime
import csv
import io

from database import add_case, UPLOADS_DIR, update_case, get_case_by_id

CASE_TYPES = ['مزایده', 'مناقصه', 'تفاهم نامه', 'صورت جلسات', 'آموزشی کارگاهی', 'اجاره سالن ها', 'اجاره ورزشی', 'نانوایی', 'بوفه', 'مرکز رشد', 'مشاوره ای', 'پژوهشی', 'رستوران', 'خوابگاه', 'آرایشگاه', 'اماکن مازاد', 'سایر']

# Try to import PIL, handle if not found
try:
    from PIL import Image
    import openpyxl
    from openpyxl.styles import Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Import Error: {e}. Please install required library (Pillow).")

PERSIAN_TO_ENGLISH_MAP = {
    '۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4',
    '۵': '5', '۶': '6', '۷': '7', '۸': '8', '۹': '9'
}

def convert_persian_to_english(text):
    """Converts Persian/Arabic numerals in a string to English numerals."""
    for p, e in PERSIAN_TO_ENGLISH_MAP.items():
        text = text.replace(p, e)
    return text

def convert_english_to_persian(text):
    """Converts English numerals in a string to Persian numerals."""
    english_to_persian_map = {v: k for k, v in PERSIAN_TO_ENGLISH_MAP.items()}
    for e, p in english_to_persian_map.items():
        text = text.replace(e, p)
    return text

class CustomJalaliCalendar(ctk.CTkFrame):
    """A custom Jalali calendar widget built with customtkinter."""
    def __init__(self, master, on_select=None, year=None, month=None, day=None):
        super().__init__(master)
        self.on_select = on_select

        today = jdatetime.date.today()
        self.year = year if year else today.year
        self.month = month if month else today.month
        self.day = day if day else today.day

        self.month_names = [
            "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
            "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
        ]
        self.weekday_names = ["ش", "ی", "د", "س", "چ", "پ", "ج"]

        # --- Header: Month/Year selection ---
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(pady=5, fill='x')
        header_frame.grid_columnconfigure(1, weight=1)
        header_frame.grid_columnconfigure(2, weight=1)

        btn_prev = ctk.CTkButton(header_frame, text=">", width=30, command=self.prev_month)
        btn_prev.grid(row=0, column=0, padx=5)

        self.year_combo = ctk.CTkComboBox(header_frame, values=[str(y) for y in range(today.year - 10, today.year + 10)], command=self.update_calendar)
        self.year_combo.set(str(self.year))
        self.year_combo.grid(row=0, column=1, padx=5, sticky='ew')

        self.month_combo = ctk.CTkComboBox(header_frame, values=self.month_names, command=self.update_calendar)
        self.month_combo.set(self.month_names[self.month - 1])
        self.month_combo.grid(row=0, column=2, padx=5, sticky='ew')

        btn_next = ctk.CTkButton(header_frame, text="<", width=30, command=self.next_month)
        btn_next.grid(row=0, column=3, padx=5)

        # --- Calendar Body ---
        self.calendar_frame = ctk.CTkFrame(self)
        self.calendar_frame.pack(expand=True, fill="both", padx=5, pady=5)

        self.create_calendar()

    def create_calendar(self):
        # Clear previous widgets
        for widget in self.calendar_frame.winfo_children():
            widget.destroy()

        # Weekday headers
        for i, name in enumerate(self.weekday_names):
            self.calendar_frame.grid_columnconfigure(i, weight=1)
            lbl = ctk.CTkLabel(self.calendar_frame, text=name, font=('vazirmatn', 12, 'bold'))
            lbl.grid(row=0, column=i, sticky='nsew')

        # Get calendar data
        first_day_of_month = jdatetime.date(self.year, self.month, 1)
        start_weekday = first_day_of_month.weekday()  # jdatetime: 0=Saturday, 6=Friday. Matches our grid.
        days_in_month = jdatetime.j_days_in_month[self.month - 1]
        if self.month == 12 and not jdatetime.date(self.year, 1, 1).isleap():
            days_in_month = 29

        day_num = 1
        for r in range(1, 7): # Max 6 rows
            self.calendar_frame.grid_rowconfigure(r, weight=1)
            for c in range(7):
                if (r == 1 and c < start_weekday) or day_num > days_in_month:
                    continue
                
                btn = ctk.CTkButton(
                    self.calendar_frame,
                    text=str(day_num),
                    command=lambda d=day_num: self.select_date(d)
                )
                btn.grid(row=r, column=c, sticky='nsew', padx=1, pady=1)
                day_num += 1

    def update_calendar(self, event=None):
        self.year = int(self.year_combo.get())
        self.month = self.month_names.index(self.month_combo.get()) + 1
        self.create_calendar()

    def next_month(self):
        self.month += 1
        if self.month > 12:
            self.month = 1
            self.year += 1
        self.year_combo.set(str(self.year))
        self.month_combo.set(self.month_names[self.month - 1])
        self.update_calendar()

    def prev_month(self):
        self.month -= 1
        if self.month < 1:
            self.month = 12
            self.year -= 1
        self.year_combo.set(str(self.year))
        self.month_combo.set(self.month_names[self.month - 1])
        self.update_calendar()

    def select_date(self, day):
        if self.on_select:
            date_str = f"{self.year}-{self.month:02d}-{day:02d}"
            self.on_select(date_str)

def open_calendar(master, entry_widget):
    """Opens a Jalali calendar to select a date and inserts it into the entry_widget."""
    def on_date_select(date_str):
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, date_str)
        cal_win.destroy()
        entry_widget.event_generate('<FocusOut>')

    cal_win = ctk.CTkToplevel(master)
    cal_win.title("انتخاب تاریخ")
    cal_win.geometry("400x350")
    cal_win.resizable(False, False)

    try:
        current_date_str = entry_widget.get()
        year, month, day = map(int, current_date_str.split('-'))
        cal = CustomJalaliCalendar(cal_win, on_select=on_date_select, year=year, month=month, day=day)
    except (ValueError, IndexError):
        cal = CustomJalaliCalendar(cal_win, on_select=on_date_select)

    cal.pack(pady=10, padx=10, fill="both", expand=True)

def open_add_record(master):
    """Open the Add Record window using customtkinter.

    master: parent window (ctk.CTk or ctk.CTkToplevel)
    """
    master.withdraw()
    top = ctk.CTkToplevel(master)
    top.title('ایجاد پرونده جدید — سیستم مدیریت پرونده مهاجر')
    top.geometry('720x550')

    def on_close():
        master.deiconify()
        top.destroy()
    
    top.protocol("WM_DELETE_WINDOW", on_close)
    
    # Configure grid layout: 3 columns (spacer, main content, labels)
    top.grid_columnconfigure(0, weight=0)
    top.grid_columnconfigure(1, weight=1)
    top.grid_columnconfigure(2, weight=0)

    selected_files = {'list': []}
    pad = 8
    font_spec = ('vazirmatn', 13, 'bold')
    font_entry = ('vazirmatn', 13)  # Regular font for entry fields

    # --- Row 0: Title and Case Type ---
    # Create a frame to hold case type and title side-by-side
    title_case_frame = ctk.CTkFrame(top)
    title_case_frame.grid(row=0, column=1, columnspan=2, sticky='ew', padx=pad, pady=pad)
    title_case_frame.grid_columnconfigure(0, weight=1)  # Case type combobox expands
    title_case_frame.grid_columnconfigure(1, weight=0)  # Case type label
    title_case_frame.grid_columnconfigure(2, weight=1)  # Title entry expands
    title_case_frame.grid_columnconfigure(3, weight=0)  # Title label
    
    # Case type label and combobox (balanced width)
    lbl_case_type = ctk.CTkLabel(title_case_frame, text='نوع پرونده', font=font_spec)
    lbl_case_type.grid(row=0, column=1, sticky='e', padx=(0, 4))
    combo_case_type = ctk.CTkComboBox(title_case_frame, values=CASE_TYPES, font=font_spec, width=150)
    combo_case_type.set(CASE_TYPES[0])
    combo_case_type.grid(row=0, column=0, sticky='ew', padx=(0, pad))
    entry_case_type_other = ctk.CTkEntry(title_case_frame, justify='right', font=font_entry, width=150)
    entry_case_type_other.grid(row=0, column=0, sticky='ew', padx=(0, pad))
    entry_case_type_other.grid_remove()

    # Title label and entry
    lbl_title = ctk.CTkLabel(title_case_frame, text='عنوان پرونده', font=font_spec)
    lbl_title.grid(row=0, column=3, sticky='e', padx=(0, 4))
    entry_title = ctk.CTkEntry(title_case_frame, justify='right', font=font_entry)
    entry_title.grid(row=0, column=2, sticky='ew', padx=(0, pad))
    
    def on_case_type_change(choice):
        if choice == 'سایر':
            combo_case_type.grid_remove()
            entry_case_type_other.grid()
        else:
            entry_case_type_other.grid_remove()
            combo_case_type.grid()

    combo_case_type.configure(command=on_case_type_change)

    # --- Row 1: Date and Subject (2-column layout) ---
    date_subject_frame = ctk.CTkFrame(top)
    date_subject_frame.grid(row=1, column=1, columnspan=2, sticky='ew', padx=pad, pady=pad)
    date_subject_frame.grid_columnconfigure(0, weight=1) # Subject expands
    date_subject_frame.grid_columnconfigure(1, weight=0) # Subject label
    date_subject_frame.grid_columnconfigure(2, weight=0) # Date entry
    date_subject_frame.grid_columnconfigure(3, weight=0) # Calendar button
    date_subject_frame.grid_columnconfigure(4, weight=0) # Today button
    date_subject_frame.grid_columnconfigure(5, weight=0) # Date label

    # RTL: Date on right, Subject on left
    lbl_date = ctk.CTkLabel(date_subject_frame, text='تاریخ پرونده', font=font_spec)
    lbl_date.grid(row=0, column=5, sticky='e', padx=(0, 4))

    def normalize_date_input(event):
        widget = event.widget
        current_text = widget.get()
        new_text = convert_persian_to_english(current_text)
        # Also trigger duration computation on key release
        compute_duration_label()
        if current_text != new_text:
            widget.delete(0, tk.END)
            widget.insert(0, new_text)

    # Today button (between label and entry)
    btn_today = ctk.CTkButton(date_subject_frame, text='امروز', command=lambda: (entry_date.delete(0, tk.END), entry_date.insert(0, jdatetime.date.today().strftime('%Y-%m-%d'))), font=('vazirmatn', 11, 'bold'), width=80)
    btn_today.grid(row=0, column=4, sticky='e', padx=(0, 4))

    btn_cal_date = ctk.CTkButton(date_subject_frame, text='تقویم', command=lambda: open_calendar(top, entry_date), width=50, font=('vazirmatn', 11, 'bold'))
    btn_cal_date.grid(row=0, column=3, sticky='e', padx=(0, 4))

    entry_date = ctk.CTkEntry(date_subject_frame, justify='right', font=font_entry, width=150)
    entry_date.grid(row=0, column=2, sticky='e', padx=(0, pad))
    
    def set_today():
        entry_date.delete(0, tk.END)
        entry_date.insert(0, jdatetime.date.today().strftime('%Y-%m-%d'))
    set_today()
    entry_date.bind('<KeyRelease>', normalize_date_input)

    lbl_subject = ctk.CTkLabel(date_subject_frame, text='موضوع', font=font_spec)
    lbl_subject.grid(row=0, column=1, sticky='e', padx=(pad, 4))
    entry_subject = ctk.CTkEntry(date_subject_frame, justify='right', font=font_entry)
    entry_subject.grid(row=0, column=0, sticky='ew', padx=(0, pad))

    # --- Row 2: Parties (2x2 grid) ---
    parties_frame = ctk.CTkFrame(top)
    parties_frame.grid(row=2, column=1, columnspan=2, sticky='ew', padx=pad, pady=pad)
    parties_frame.grid_columnconfigure(0, weight=1)  # Right column expands
    parties_frame.grid_columnconfigure(2, weight=1)  # Left column expands

    # Row 0 of parties frame: موجر (Owner) and مستاجر (Tenant)
    lbl_mojer = ctk.CTkLabel(parties_frame, text='موجر', font=font_spec)
    lbl_mojer.grid(row=0, column=3, sticky='e')
    entry_mojer = ctk.CTkEntry(parties_frame, justify='right', font=font_entry)
    entry_mojer.grid(row=0, column=2, sticky='ew', padx=(0, pad))

    lbl_mostajjer = ctk.CTkLabel(parties_frame, text='مستاجر', font=font_spec)
    lbl_mostajjer.grid(row=0, column=1, sticky='e', padx=(0, 4))
    entry_mostajjer = ctk.CTkEntry(parties_frame, justify='right', font=font_entry)
    entry_mostajjer.grid(row=0, column=0, sticky='ew', padx=(0, pad))

    # Row 1 of parties frame: کارفرما (Employer) and پیمانکار (Contractor)
    lbl_karfarma = ctk.CTkLabel(parties_frame, text='کارفرما', font=font_spec)
    lbl_karfarma.grid(row=1, column=3, sticky='e', pady=(pad, 0))
    entry_karfarma = ctk.CTkEntry(parties_frame, justify='right', font=font_entry)
    entry_karfarma.grid(row=1, column=2, sticky='ew', padx=(0, pad), pady=(pad, 0))

    lbl_piman = ctk.CTkLabel(parties_frame, text='پیمانکار', font=font_spec)
    lbl_piman.grid(row=1, column=1, sticky='e', padx=(0, 4), pady=(pad, 0))
    entry_piman = ctk.CTkEntry(parties_frame, justify='right', font=font_entry)
    entry_piman.grid(row=1, column=0, sticky='ew', padx=(0, pad), pady=(pad, 0))

    # Bank account data storage
    bank_data = {}
    
    # Guarantee data storage
    guarantee_data = {}

    def open_bank_account_window():
        """Open a small window for entering bank account details."""
        bank_win = ctk.CTkToplevel(top)
        bank_win.title('افزودن اطلاعات حساب بانکی')
        bank_win.geometry('400x360')
        bank_win.resizable(False, False)

        # Configure grid layout for RTL
        bank_win.grid_columnconfigure(0, weight=0)
        bank_win.grid_columnconfigure(1, weight=1)

        font_bank = ('vazirmatn', 12, 'bold')

        # نام صاحب حساب (Account owner name)
        lbl_owner = ctk.CTkLabel(bank_win, text='نام صاحب حساب', font=font_bank)
        lbl_owner.grid(row=0, column=1, sticky='e', padx=(pad, pad), pady=(pad, 0))
        entry_owner = ctk.CTkEntry(bank_win, justify='right', font=('vazirmatn', 12))
        entry_owner.grid(row=0, column=0, sticky='ew', padx=(pad, pad), pady=(pad, 0))
        entry_owner.insert(0, bank_data.get('owner_name', ''))

        # شماره حساب (Account number)
        lbl_account = ctk.CTkLabel(bank_win, text='شماره حساب', font=font_bank)
        lbl_account.grid(row=1, column=1, sticky='e', padx=(pad, pad), pady=pad)
        entry_account = ctk.CTkEntry(bank_win, justify='right', font=('vazirmatn', 12))
        entry_account.grid(row=1, column=0, sticky='ew', padx=(pad, pad), pady=pad)
        entry_account.insert(0, bank_data.get('account_number', ''))

        # شماره شبا (SHABA number)
        lbl_shaba = ctk.CTkLabel(bank_win, text='شماره شبا', font=font_bank)
        lbl_shaba.grid(row=2, column=1, sticky='e', padx=(pad, pad), pady=pad)
        entry_shaba = ctk.CTkEntry(bank_win, justify='right', font=('vazirmatn', 12))
        entry_shaba.grid(row=2, column=0, sticky='ew', padx=(pad, pad), pady=pad) # Corrected row
        entry_shaba.insert(0, bank_data.get('shaba_number', ''))

        # نام بانک (Bank Name)
        lbl_bank_name = ctk.CTkLabel(bank_win, text='نام بانک', font=font_bank)
        lbl_bank_name.grid(row=3, column=1, sticky='e', padx=(pad, pad), pady=pad)
        entry_bank_name = ctk.CTkEntry(bank_win, justify='right', font=('vazirmatn', 12))
        entry_bank_name.grid(row=3, column=0, sticky='ew', padx=(pad, pad), pady=pad)
        entry_bank_name.insert(0, bank_data.get('bank_name', ''))

        # شعبه بانک (Bank Branch)
        lbl_branch = ctk.CTkLabel(bank_win, text='شعبه بانک', font=font_bank)
        lbl_branch.grid(row=4, column=1, sticky='e', padx=(pad, pad), pady=pad)
        entry_branch = ctk.CTkEntry(bank_win, justify='right', font=('vazirmatn', 12))
        entry_branch.grid(row=4, column=0, sticky='ew', padx=(pad, pad), pady=pad)
        entry_branch.insert(0, bank_data.get('bank_branch', ''))

        # شناسه پرداخت (Payment ID)
        lbl_payment_id = ctk.CTkLabel(bank_win, text='شناسه پرداخت', font=font_bank)
        lbl_payment_id.grid(row=5, column=1, sticky='e', padx=(pad, pad), pady=pad)
        entry_payment_id = ctk.CTkEntry(bank_win, justify='right', font=('vazirmatn', 12))
        entry_payment_id.grid(row=5, column=0, sticky='ew', padx=(pad, pad), pady=pad)
        entry_payment_id.insert(0, bank_data.get('payment_id', ''))

        # شماره کارت (Card number)
        lbl_card = ctk.CTkLabel(bank_win, text='شماره کارت', font=font_bank)
        lbl_card.grid(row=6, column=1, sticky='e', padx=(pad, pad), pady=pad)
        entry_card = ctk.CTkEntry(bank_win, justify='right', font=('vazirmatn', 12))
        entry_card.grid(row=6, column=0, sticky='ew', padx=(pad, pad), pady=pad)
        entry_card.insert(0, bank_data.get('card_number', ''))

        # Save button
        def save_bank_data():
            bank_data['owner_name'] = entry_owner.get().strip()
            bank_data['account_number'] = entry_account.get().strip()
            bank_data['shaba_number'] = entry_shaba.get().strip()
            bank_data['bank_name'] = entry_bank_name.get().strip()
            bank_data['bank_branch'] = entry_branch.get().strip()
            bank_data['payment_id'] = entry_payment_id.get().strip()
            bank_data['card_number'] = entry_card.get().strip()
            bank_win.destroy()

        btn_save_bank = ctk.CTkButton(bank_win, text='ذخیره', command=save_bank_data, font=('vazirmatn', 12, 'bold'))
        btn_save_bank.grid(row=7, column=0, columnspan=2, pady=(pad, pad), padx=pad, sticky='ew')

    def open_guarantee_window():
        """Open a window for entering contract guarantee details."""
        guarantee_win = ctk.CTkToplevel(top)
        guarantee_win.title('افزودن تضمین قرارداد')
        guarantee_win.geometry('350x180')
        guarantee_win.resizable(False, False)

        guarantee_win.grid_columnconfigure(1, weight=1)
        font_guarantee = ('vazirmatn', 12, 'bold')
        
        # Guarantee Amount
        lbl_amount_guarantee = ctk.CTkLabel(guarantee_win, text='مبلغ تضمین', font=font_guarantee)
        lbl_amount_guarantee.grid(row=0, column=3, sticky='e', padx=pad, pady=pad)
        lbl_rial_guarantee = ctk.CTkLabel(guarantee_win, text='ریال', font=font_guarantee)
        lbl_rial_guarantee.grid(row=0, column=0, sticky='w', padx=(4, pad))

        def format_guarantee_amount(event=None):
            text = entry_amount.get().replace(',', '')
            if text.isdigit():
                formatted_text = f"{int(text):,}"
                entry_amount.delete(0, tk.END)
                entry_amount.insert(0, formatted_text)

        vcmd_guarantee = (guarantee_win.register(validate_amount), '%P')
        entry_amount = ctk.CTkEntry(guarantee_win, justify='right', font=('vazirmatn', 12), validate='key', validatecommand=vcmd_guarantee)
        entry_amount.grid(row=0, column=1, columnspan=2, sticky='ew', padx=pad, pady=pad)
        entry_amount.insert(0, guarantee_data.get('guarantee_amount', '').replace(',', ''))
        entry_amount.bind("<KeyRelease>", format_guarantee_amount)
        format_guarantee_amount()

        # Guarantee Type
        lbl_type = ctk.CTkLabel(guarantee_win, text='نوع ضمانت', font=font_guarantee)
        lbl_type.grid(row=1, column=3, sticky='e', padx=pad, pady=pad)
        guarantee_types = ['چک', 'سفته', 'ضمانت نامه بانکی', 'تعهد', 'سایر']
        combo_type = ctk.CTkComboBox(guarantee_win, values=guarantee_types, font=('vazirmatn', 12), state='readonly')
        combo_type.grid(row=1, column=0, columnspan=3, sticky='ew', padx=pad, pady=pad)
        entry_type_other = ctk.CTkEntry(guarantee_win, justify='right', font=('vazirmatn', 12))
        entry_type_other.grid(row=1, column=0, columnspan=3, sticky='ew', padx=pad, pady=pad)

        # Initialize state
        gt_val = guarantee_data.get('guarantee_type', '')
        if gt_val in guarantee_types and gt_val != 'سایر':
            combo_type.set(gt_val)
            entry_type_other.grid_remove()
        else:
            combo_type.set('سایر' if gt_val else guarantee_types[0])
            entry_type_other.insert(0, gt_val)
            if combo_type.get() != 'سایر':
                entry_type_other.grid_remove()
            else:
                combo_type.grid_remove()

        def on_guarantee_type_change(event=None):
            if combo_type.get() == 'سایر':
                combo_type.grid_remove()
                entry_type_other.grid()
            else:
                entry_type_other.grid_remove()
                combo_type.grid()
        
        combo_type.configure(command=on_guarantee_type_change)

        def save_guarantee_data():
            guarantee_data['guarantee_amount'] = entry_amount.get().strip().replace(',', '')
            guarantee_data['guarantee_type'] = entry_type_other.get().strip() if combo_type.get() == 'سایر' else combo_type.get()
            guarantee_win.destroy()

        btn_save = ctk.CTkButton(guarantee_win, text='ذخیره', command=save_guarantee_data, font=font_guarantee)
        btn_save.grid(row=2, column=0, columnspan=3, sticky='ew', padx=pad, pady=pad)

    # --- Row 3: Contract Amount and Bank Account Button ---
    amount_frame = ctk.CTkFrame(top)
    amount_frame.grid(row=3, column=1, columnspan=2, sticky='ew', padx=pad, pady=pad)
    amount_frame.grid_columnconfigure(0, weight=1)
    amount_frame.grid_columnconfigure(2, weight=0)

    # A sub-frame for the two buttons
    buttons_frame = ctk.CTkFrame(amount_frame, fg_color="transparent")
    buttons_frame.grid(row=0, column=0, sticky='ew', padx=(0, pad))
    buttons_frame.grid_columnconfigure(0, weight=1)
    buttons_frame.grid_columnconfigure(1, weight=1)

    btn_guarantee = ctk.CTkButton(buttons_frame, text='تضمین قرارداد', command=open_guarantee_window, font=font_spec)
    btn_guarantee.grid(row=0, column=0, sticky='ew', padx=(0, 2))

    btn_bank = ctk.CTkButton(buttons_frame, text='افزودن حساب', command=open_bank_account_window, font=font_spec)
    btn_bank.grid(row=0, column=1, sticky='ew', padx=(2, 0))

    lbl_rial = ctk.CTkLabel(amount_frame, text='ریال', font=('vazirmatn', 12, 'bold'))
    lbl_rial.grid(row=0, column=1, sticky='w', padx=(0, 4))

    lbl_contract_amount = ctk.CTkLabel(amount_frame, text='مبلغ قرارداد', font=font_spec)
    lbl_contract_amount.grid(row=0, column=3, sticky='e', padx=(0, pad))
    
    def format_amount(event=None):
        text = entry_contract_amount.get().replace(',', '')
        if text.isdigit():
            formatted_text = f"{int(text):,}"
            entry_contract_amount.delete(0, tk.END)
            entry_contract_amount.insert(0, formatted_text)

    def validate_amount(P):
        return str.isdigit(convert_persian_to_english(P).replace(',', '')) or P == ""

    vcmd = (top.register(validate_amount), '%P')
    entry_contract_amount = ctk.CTkEntry(amount_frame, justify='right', font=font_entry, validate='key', validatecommand=vcmd)
    entry_contract_amount.grid(row=0, column=2, sticky='ew', padx=(0, pad))
    entry_contract_amount.bind("<KeyRelease>", format_amount)

    # --- Row 4: Duration (از - تا) ---
    # Dedicated row for duration with RTL layout: [calc_label] [تا_entry]
    duration_frame = ctk.CTkFrame(top)
    duration_frame.grid(row=4, column=1, columnspan=2, sticky='ew', padx=pad, pady=pad)
    duration_frame.grid_columnconfigure(0, weight=1) # calc label
    duration_frame.grid_columnconfigure(1, weight=1) # to entry
    duration_frame.grid_columnconfigure(2, weight=0) # to cal button
    duration_frame.grid_columnconfigure(3, weight=0) # to label
    duration_frame.grid_columnconfigure(4, weight=1) # from entry
    duration_frame.grid_columnconfigure(5, weight=0) # from cal button
    duration_frame.grid_columnconfigure(6, weight=0) # from label

    # RTL layout: from right to left (columns 4, 3, 2, 1, 0)
    lbl_duration_from = ctk.CTkLabel(duration_frame, text='از:', font=font_spec)
    lbl_duration_from.grid(row=0, column=6, sticky='e', padx=(0, 4))

    btn_cal_from = ctk.CTkButton(duration_frame, text='تقویم', command=lambda: open_calendar(top, entry_duration_from), width=60, font=('vazirmatn', 11, 'bold'))
    btn_cal_from.grid(row=0, column=5, sticky='e', padx=(0, 4))

    entry_duration_from = ctk.CTkEntry(duration_frame, justify='right', font=font_entry, width=150)    
    entry_duration_from.grid(row=0, column=4, sticky='ew', padx=(0, pad))
    entry_duration_from.insert(0, jdatetime.date.today().strftime('%Y-%m-%d'))
    entry_duration_from.bind('<KeyRelease>', normalize_date_input)

    lbl_duration_to = ctk.CTkLabel(duration_frame, text='تا:', font=font_spec)
    lbl_duration_to.grid(row=0, column=3, sticky='e', padx=(0, 4))

    entry_duration_to = ctk.CTkEntry(duration_frame, justify='right', font=font_entry, width=150)
    entry_duration_to.grid(row=0, column=1, sticky='ew', padx=(0, pad))
    entry_duration_to.insert(0, jdatetime.date.today().strftime('%Y-%m-%d'))
    entry_duration_to.bind('<KeyRelease>', normalize_date_input)

    lbl_duration_calc = ctk.CTkLabel(duration_frame, text='', font=('vazirmatn', 12), text_color='#0066cc')
    lbl_duration_calc.grid(row=0, column=0, sticky='w', padx=(pad, 0))

    btn_cal_to = ctk.CTkButton(duration_frame, text='تقویم', command=lambda: open_calendar(top, entry_duration_to), width=60, font=('vazirmatn', 11, 'bold'))
    btn_cal_to.grid(row=0, column=2, sticky='e', padx=(0, 4))

    def compute_duration_label():
        from_val = entry_duration_from.get().strip()
        to_val = entry_duration_to.get().strip()
        if not from_val or not to_val:
            lbl_duration_calc.configure(text='')
            return
        try:
            d1 = jdatetime.datetime.strptime(from_val, '%Y-%m-%d').date()
            d2 = jdatetime.datetime.strptime(to_val, '%Y-%m-%d').date()
            today = jdatetime.date.today()
            
            # Check if end date is past
            if d2 < today:
                status_var.set('راکد')
            else:
                status_var.set('در جریان')
            
            # convert to Gregorian for timedelta
            g1 = d1.togregorian()
            g2 = d2.togregorian()
            diff_days = (g2 - g1).days
            if diff_days < 0:
                lbl_duration_calc.configure(text='تاریخ پایان قبل از شروع است')
            else:
                # Calculate years, months, days
                years = d2.year - d1.year
                months = d2.month - d1.month
                days = d2.day - d1.day
                if days < 0:
                    months -= 1
                    # get days in previous month
                    if d2.month == 1:
                        prev_month_days = 31
                    else:
                        prev_month_date = jdatetime.date(d2.year, d2.month - 1, 1)
                        prev_month_days = 31 if d2.month - 1 in [1, 3, 5, 7, 9, 11] else (30 if d2.month - 1 in [4, 6, 8, 10] else 29)
                    days += prev_month_days
                if months < 0:
                    years -= 1
                    months += 12
                lbl_duration_calc.configure(text=f'مدت قرارداد: {years} سال، {months} ماه، {days} روز')
        except Exception as e:
            lbl_duration_calc.configure(text='فرمت تاریخ باید YYYY-MM-DD باشد')

    # bind compute on change (focus out)
    entry_duration_from.bind('<FocusOut>', lambda e: compute_duration_label())
    entry_duration_to.bind('<FocusOut>', lambda e: compute_duration_label())
    # call once at start to show initial duration
    compute_duration_label()

    # --- Row 5: Description ---
    lbl_desc = ctk.CTkLabel(top, text='توضیحات', font=font_spec)
    lbl_desc.grid(row=5, column=2, sticky='ne', padx=(0, pad), pady=(pad, 0))
    text_desc = ctk.CTkTextbox(top, height=150, font=font_entry)
    # text_desc.tag_configure("right", justify="right") # Not supported in ctk
    text_desc.grid(row=5, column=1, sticky='nsew', padx=pad, pady=(0, pad))
    top.grid_rowconfigure(5, weight=1) # Allow textbox to expand vertically

    combo_case_type.configure(command=on_case_type_change)

    # --- Row 6: Attachments + Export Checkbox + Status ---
    files_frame = ctk.CTkFrame(top)
    files_frame.grid(row=6, column=0, columnspan=3, sticky='ew', padx=pad, pady=pad)
    files_frame.grid_columnconfigure(0, weight=1)  # File label expands to left
    files_frame.grid_columnconfigure(1, weight=0)  # Attach button
    files_frame.grid_columnconfigure(2, weight=0)  # Export checkbox
    files_frame.grid_columnconfigure(3, weight=0)  # Status radios
    files_frame.grid_columnconfigure(4, weight=0)  # Button stays small on right

    lbl_files = ctk.CTkLabel(files_frame, text='فایلی انتخاب نشده است', font=('vazirmatn', 12), text_color="gray")
    lbl_files.grid(row=0, column=0, sticky='w', padx=(pad, 0))

    def select_files():
        files = filedialog.askopenfilenames(title='انتخاب فایل‌ها')
        if files:
            selected_files['list'] = list(files)
            lbl_files.configure(text=f'{len(selected_files["list"])} فایل انتخاب شد')

    btn_attach = ctk.CTkButton(files_frame, text='افزودن فایل پیوست', command=select_files, font=('vazirmatn', 13, 'bold'))
    btn_attach.grid(row=0, column=1, sticky='e', padx=(pad, 0))

    # Export checkbox (default ON)
    export_var = tk.BooleanVar(value=True)
    chk_export = ctk.CTkCheckBox(
        files_frame,
        text='ایجاد فایل XLSX',
        variable=export_var,
        font=('vazirmatn', 11)
    )
    chk_export.grid(row=0, column=2, sticky='e', padx=pad)

    # Status indicator (Radios)
    status_var = tk.StringVar(value='در جریان')
    
    frame_status = ctk.CTkFrame(files_frame, fg_color="transparent")
    frame_status.grid(row=0, column=3, sticky='e', padx=pad)
    
    lbl_status = ctk.CTkLabel(frame_status, text='وضعیت:', font=('vazirmatn', 11, 'bold'))
    lbl_status.pack(side='right', padx=(0, 8))
    
    rb_active = ctk.CTkRadioButton(frame_status, text='در جریان', variable=status_var, value='در جریان', font=('vazirmatn', 11), text_color='green')
    rb_active.pack(side='right', padx=2)
    
    rb_inactive = ctk.CTkRadioButton(frame_status, text='راکد', variable=status_var, value='راکد', font=('vazirmatn', 11), text_color='red')
    rb_inactive.pack(side='right', padx=2)

    # --- Row 7: Save Button ---
    btn_save = ctk.CTkButton(top, text='ذخیره پرونده', command=lambda: save_case(top, selected_files, entry_title, entry_date, entry_duration_from, entry_duration_to, entry_mojer, entry_mostajjer, entry_karfarma, entry_piman, entry_subject, text_desc, entry_contract_amount, combo_case_type, entry_case_type_other, bank_data, status_var, export_var, guarantee_data), font=('vazirmatn', 14, 'bold'), height=40)
    btn_save.grid(row=7, column=0, columnspan=3, sticky='ew', padx=pad, pady=(pad, pad))


def save_case(top, selected_files, entry_title, entry_date, entry_duration_from, entry_duration_to, entry_mojer, entry_mostajjer, entry_karfarma, entry_piman, entry_subject, text_desc, entry_contract_amount, combo_case_type, entry_case_type_other, bank_data, status_var, export_var=None, guarantee_data=None):
    title = entry_title.get().strip()
    if not title:
        messagebox.showwarning('خطا', 'عنوان پرونده الزامی است', parent=top)
        return

    now = jdatetime.datetime.now()
    case_id = now.strftime('%Y%m%d%H%M%S')

    date_str = entry_date.get().strip() or jdatetime.date.today().strftime('%Y-%m-%d')

    folder = os.path.join(UPLOADS_DIR, case_id)
    os.makedirs(folder, exist_ok=True)

    for f in selected_files['list']:
        try:
            shutil.copy2(f, folder)
        except Exception:
            pass

    # compute duration string
    from_val = entry_duration_from.get().strip()
    to_val = entry_duration_to.get().strip()
    duration_str = ''
    try:
        if from_val and to_val:
            d1 = jdatetime.datetime.strptime(from_val, '%Y-%m-%d').date()
            d2 = jdatetime.datetime.strptime(to_val, '%Y-%m-%d').date()
            diff_days = (d2.togregorian() - d1.togregorian()).days
            if diff_days >= 0:
                duration_str = f'{convert_english_to_persian(str(diff_days))} \u200eروز'
            else:
                duration_str = ''
    except Exception:
        duration_str = ''

    # case type
    case_type_val = combo_case_type.get()
    if case_type_val == 'سایر':
        case_type_val = entry_case_type_other.get().strip()

    data = {
        'id': case_id,
        'title': title,
        'date': date_str,
        'duration': duration_str,
        'duration_from': from_val,
        'duration_to': to_val,
        'mojer': entry_mojer.get().strip(),
        'mostajjer': entry_mostajjer.get().strip(),
        'karfarma': entry_karfarma.get().strip(),
        'piman': entry_piman.get().strip(),
        'subject': entry_subject.get().strip(),
        'contract_amount': entry_contract_amount.get().strip().replace(',', ''),
        'bank_owner_name': bank_data.get('owner_name', '').strip(),
        'bank_account_number': bank_data.get('account_number', '').strip(),
        'bank_shaba_number': bank_data.get('shaba_number', '').strip(),
        'bank_card_number': bank_data.get('card_number', '').strip(),
        'bank_name': bank_data.get('bank_name', '').strip(),
        'bank_branch': bank_data.get('bank_branch', '').strip(),
        'payment_id': bank_data.get('payment_id', '').strip(),        
        'guarantee_amount': guarantee_data.get('guarantee_amount', '').strip() if guarantee_data else '',
        'guarantee_type': guarantee_data.get('guarantee_type', '').strip() if guarantee_data else '',
        'description': text_desc.get('1.0', tk.END).strip(),
        'folder_path': folder,
        'case_type': case_type_val,
        'status': status_var.get()
    }

    # Export to CSV and XLSX if checkbox is enabled
    if export_var and export_var.get():
        export_case_to_files(data)

    try:
        add_case(data)
        messagebox.showinfo('موفق', f'پرونده با شماره {case_id} ایجاد شد', parent=top)
        top.master.deiconify()
        top.destroy()
    except Exception as e:
        messagebox.showerror('خطا', f'خطا در ذخیره: {e}', parent=top)


def export_case_to_files(data):
    """Export case data to CSV and TXT files in the case folder."""
    try:
        case_id = data['id']
        folder = data['folder_path']
        
        # Format amounts with ریال
        contract_amount_display = f"{data['contract_amount']} ریال" if data['contract_amount'] else ''
        guarantee_amount_display = f"{data.get('guarantee_amount', '')} ریال" if data.get('guarantee_amount', '') else ''
        
        # Prepare case data as key-value pairs
        case_info = [
            ('شناسه بایگانی', case_id),
            ('عنوان', data['title']),
            ('تاریخ پرونده', data['date']),
            ('موضوع', data['subject']),
            ('نوع پرونده', data['case_type']),
            ('مدت قرارداد', data['duration']),
            ('تاریخ شروع', data['duration_from']),
            ('تاریخ پایان', data['duration_to']),
            ('موجر', data['mojer']),
            ('مستاجر', data['mostajjer']),
            ('کارفرما', data['karfarma']),
            ('پیمانکار', data.get('piman', '')),
            ('مبلغ قرارداد', contract_amount_display),
            ('نام صاحب حساب', data['bank_owner_name']),
            ('شماره حساب', data['bank_account_number']),
            ('شماره شبا', data['bank_shaba_number']),
            ('شماره کارت', data['bank_card_number']),
            ('نام بانک', data.get('bank_name', '')),
            ('شعبه بانک', data.get('bank_branch', '')),
            ('شناسه پرداخت', data.get('payment_id', '')),
            ('مبلغ تضمین', guarantee_amount_display),
            ('نوع ضمانت', data.get('guarantee_type', '')),
            ('وضعیت', data['status']),
            ('توضیحات', data['description']),
        ]
        
        # Export to XLSX
        xlsx_path = os.path.join(folder, f'{case_id}.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "جزئیات پرونده"
        sheet.sheet_view.rightToLeft = True

        # --- Styles ---
        bold_font = Font(name='Vazirmatn', bold=True)
        regular_font = Font(name='Vazirmatn')
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        # readingOrder=2 forces RTL direction in the cell, which fixes display on Windows
        center_alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)

        # --- Data Preparation ---
        headers = [item[0] for item in case_info]
        values = [item[1] for item in case_info]

        # Add "Row" column
        headers.insert(0, 'ردیف')
        values.insert(0, 1)

        # --- Write to Sheet ---
        sheet.append(headers)
        sheet.append([str(v) if v is not None else '' for v in values])

        # --- Apply Styles and Adjust Widths ---
        for col_idx, col in enumerate(sheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell_idx, cell in enumerate(col):
                # Apply styles
                cell.font = bold_font if cell.row == 1 else regular_font
                cell.border = thin_border
                cell.alignment = center_alignment
                # Find max length in column
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(xlsx_path)
    except Exception as e:
        print(f"Error exporting case files: {e}")
        # Don't block case saving if export fails
        pass


def open_edit_record(master, case_id):
    """Open edit window identical to create window but prefilled with data for case_id."""
    data = get_case_by_id(case_id)
    if not data:
        messagebox.showerror('خطا', 'پرونده یافت نشد')
        return

    master.withdraw()
    top = ctk.CTkToplevel(master)
    top.title(f'ویرایش پرونده {case_id} — سیستم مدیریت پرونده مهاجر')
    top.geometry('720x550')

    def on_close():
        master.deiconify()
        top.destroy()
    
    top.protocol("WM_DELETE_WINDOW", on_close)

    # Configure grid
    top.grid_columnconfigure(0, weight=0)
    top.grid_columnconfigure(1, weight=1)
    top.grid_columnconfigure(2, weight=0)

    selected_files = {'list': []}
    pad = 8
    font_spec = ('vazirmatn', 13, 'bold')
    font_entry = ('vazirmatn', 13)  # Regular font for entry fields

    # Row 0: Title and Case Type
    title_case_frame = ctk.CTkFrame(top)
    title_case_frame.grid(row=0, column=1, columnspan=2, sticky='ew', padx=pad, pady=pad)
    title_case_frame.grid_columnconfigure(0, weight=1)  # Case type combobox expands
    title_case_frame.grid_columnconfigure(1, weight=0)  # Case type label
    title_case_frame.grid_columnconfigure(2, weight=1)  # Title entry expands
    title_case_frame.grid_columnconfigure(3, weight=0)  # Title label
    
    # Case type label and combobox (balanced width)
    lbl_case_type = ctk.CTkLabel(title_case_frame, text='نوع پرونده', font=font_spec)
    lbl_case_type.grid(row=0, column=1, sticky='e', padx=(0, 4))
    combo_case_type = ctk.CTkComboBox(title_case_frame, values=CASE_TYPES, font=font_spec, width=150)
    combo_case_type.grid(row=0, column=0, sticky='ew', padx=(0, pad))
    ct_val = data.get('case_type') or CASE_TYPES[0]
    entry_case_type_other = ctk.CTkEntry(title_case_frame, justify='right', font=font_entry, width=150)
    entry_case_type_other.grid(row=0, column=0, sticky='ew', padx=(0, pad))

    # Title label and entry
    lbl_title = ctk.CTkLabel(title_case_frame, text='عنوان پرونده', font=font_spec)
    lbl_title.grid(row=0, column=3, sticky='e', padx=(0, 4))
    entry_title = ctk.CTkEntry(title_case_frame, justify='right', font=font_entry)
    entry_title.grid(row=0, column=2, sticky='ew', padx=(0, pad))
    entry_title.insert(0, data.get('title') or '')
    
    # Row 1: Date & Subject
    date_subject_frame = ctk.CTkFrame(top)
    date_subject_frame.grid(row=1, column=1, columnspan=2, sticky='ew', padx=pad, pady=pad)
    date_subject_frame.grid_columnconfigure(0, weight=1) # Subject expands
    date_subject_frame.grid_columnconfigure(1, weight=0) # Subject label
    date_subject_frame.grid_columnconfigure(2, weight=0) # Date entry
    date_subject_frame.grid_columnconfigure(3, weight=0) # Calendar button
    date_subject_frame.grid_columnconfigure(4, weight=0) # Today button
    date_subject_frame.grid_columnconfigure(5, weight=0) # Date label

    def normalize_date_input(event):
        widget = event.widget
        current_text = widget.get()
        new_text = convert_persian_to_english(current_text)
        # Also trigger duration computation on key release
        compute_duration_label_edit()
        if current_text != new_text:
            widget.delete(0, tk.END)
            widget.insert(0, new_text)

    # RTL: Date on right, Subject on left
    lbl_date = ctk.CTkLabel(date_subject_frame, text='تاریخ پرونده', font=font_spec)
    lbl_date.grid(row=0, column=5, sticky='e', padx=(0, 4))

    # Today button (between label and entry)
    btn_today = ctk.CTkButton(date_subject_frame, text='امروز', command=lambda: (entry_date.delete(0, tk.END), entry_date.insert(0, jdatetime.date.today().strftime('%Y-%m-%d'))), font=('vazirmatn', 11, 'bold'), width=60)
    btn_today.grid(row=0, column=4, sticky='e', padx=(0, 4))

    btn_cal_date = ctk.CTkButton(date_subject_frame, text='تقویم', command=lambda: open_calendar(top, entry_date), width=60, font=('vazirmatn', 11, 'bold'))
    btn_cal_date.grid(row=0, column=3, sticky='e', padx=(0, 4))

    entry_date = ctk.CTkEntry(date_subject_frame, justify='right', font=font_entry, width=150)
    entry_date.grid(row=0, column=2, sticky='e', padx=(0, pad))
    entry_date.insert(0, data.get('date') or jdatetime.date.today().strftime('%Y-%m-%d'))
    entry_date.bind('<KeyRelease>', normalize_date_input)

    lbl_subject = ctk.CTkLabel(date_subject_frame, text='موضوع', font=font_spec)
    lbl_subject.grid(row=0, column=1, sticky='e', padx=(pad, 4))
    entry_subject = ctk.CTkEntry(date_subject_frame, justify='right', font=font_entry)
    entry_subject.grid(row=0, column=0, sticky='ew', padx=(0, pad))
    entry_subject.insert(0, data.get('subject') or '')
    
    if ct_val in CASE_TYPES:
        combo_case_type.set(ct_val)
        entry_case_type_other.grid_remove()
    else:
        combo_case_type.set('سایر')
        combo_case_type.grid_remove()
        entry_case_type_other.insert(0, ct_val)
        entry_case_type_other.grid()

    def on_case_type_change_edit(choice):
        if choice == 'سایر':
            combo_case_type.grid_remove()
            entry_case_type_other.grid()
        else:
            entry_case_type_other.grid_remove()
            combo_case_type.grid()

    combo_case_type.configure(command=on_case_type_change_edit)

    # --- Row 2: Parties (2x2 grid) ---
    parties_frame = ctk.CTkFrame(top)
    parties_frame.grid(row=2, column=1, columnspan=2, sticky='ew', padx=pad, pady=pad)
    parties_frame.grid_columnconfigure(0, weight=1)
    parties_frame.grid_columnconfigure(2, weight=1)

    # Row 0 of parties frame: موجر (Owner) and مستاجر (Tenant)
    lbl_mojer = ctk.CTkLabel(parties_frame, text='موجر', font=font_spec)
    lbl_mojer.grid(row=0, column=3, sticky='e')
    entry_mojer = ctk.CTkEntry(parties_frame, justify='right', font=font_entry)
    entry_mojer.grid(row=0, column=2, sticky='ew', padx=(0, pad))
    entry_mojer.insert(0, data.get('mojer') or '')

    lbl_mostajjer = ctk.CTkLabel(parties_frame, text='مستاجر', font=font_spec)
    lbl_mostajjer.grid(row=0, column=1, sticky='e', padx=(0, 4))
    entry_mostajjer = ctk.CTkEntry(parties_frame, justify='right', font=font_entry)
    entry_mostajjer.grid(row=0, column=0, sticky='ew', padx=(0, pad))
    entry_mostajjer.insert(0, data.get('mostajjer') or '')

    # Row 1 of parties frame: کارفرما (Employer) and پیمانکار (Contractor)
    lbl_karfarma = ctk.CTkLabel(parties_frame, text='کارفرما', font=font_spec)
    lbl_karfarma.grid(row=1, column=3, sticky='e', pady=(pad, 0))
    entry_karfarma = ctk.CTkEntry(parties_frame, justify='right', font=font_entry)
    entry_karfarma.grid(row=1, column=2, sticky='ew', padx=(0, pad), pady=(pad, 0))
    entry_karfarma.insert(0, data.get('karfarma') or '')

    lbl_piman = ctk.CTkLabel(parties_frame, text='پیمانکار', font=font_spec)
    lbl_piman.grid(row=1, column=1, sticky='e', padx=(0, 4), pady=(pad, 0))
    entry_piman = ctk.CTkEntry(parties_frame, justify='right', font=font_entry)
    entry_piman.grid(row=1, column=0, sticky='ew', padx=(pad, pad), pady=(pad, 0))
    entry_piman.insert(0, data.get('piman') or '')

    # Bank account data storage
    bank_data = {
        'owner_name': data.get('bank_owner_name') or '',
        'account_number': data.get('bank_account_number') or '',
        'shaba_number': data.get('bank_shaba_number') or '',
        'card_number': data.get('bank_card_number') or '',
        'bank_name': data.get('bank_name') or '',
        'bank_branch': data.get('bank_branch') or '',
        'payment_id': data.get('payment_id') or ''
    }
    
    # Guarantee data storage
    guarantee_data = {
        'guarantee_amount': data.get('guarantee_amount') or '',
        'guarantee_type': data.get('guarantee_type') or ''
    }


    def open_bank_account_window():
        """Open a small window for entering bank account details."""
        bank_win = ctk.CTkToplevel(top)
        bank_win.title('ویرایش اطلاعات حساب بانکی')
        bank_win.geometry('400x360')
        bank_win.resizable(False, False)

        # Configure grid layout for RTL
        bank_win.grid_columnconfigure(0, weight=0)
        bank_win.grid_columnconfigure(1, weight=1)

        font_bank = ('vazirmatn', 12)

        # نام صاحب حساب (Account owner name)
        lbl_owner = ctk.CTkLabel(bank_win, text='نام صاحب حساب', font=('vazirmatn', 12, 'bold'))
        lbl_owner.grid(row=0, column=1, sticky='e', padx=(pad, pad), pady=(pad, 0))
        entry_owner = ctk.CTkEntry(bank_win, justify='right', font=font_bank)
        entry_owner.grid(row=0, column=0, sticky='ew', padx=(pad, pad), pady=(pad, 0))
        entry_owner.insert(0, bank_data.get('owner_name', ''))

        # شماره حساب (Account number)
        lbl_account = ctk.CTkLabel(bank_win, text='شماره حساب', font=('vazirmatn', 12, 'bold'))
        lbl_account.grid(row=1, column=1, sticky='e', padx=(pad, pad), pady=pad)
        entry_account = ctk.CTkEntry(bank_win, justify='right', font=font_bank)
        entry_account.grid(row=1, column=0, sticky='ew', padx=(pad, pad), pady=pad)
        entry_account.insert(0, bank_data.get('account_number', ''))

        # شماره شبا (SHABA number)
        lbl_shaba = ctk.CTkLabel(bank_win, text='شماره شبا', font=('vazirmatn', 12, 'bold'))
        lbl_shaba.grid(row=2, column=1, sticky='e', padx=(pad, pad), pady=pad)
        entry_shaba = ctk.CTkEntry(bank_win, justify='right', font=('vazirmatn', 12))
        entry_shaba.grid(row=2, column=0, sticky='ew', padx=(pad, pad), pady=pad)
        entry_shaba.insert(0, bank_data.get('shaba_number', ''))

        # نام بانک (Bank Name)
        lbl_bank_name = ctk.CTkLabel(bank_win, text='نام بانک', font=('vazirmatn', 12, 'bold'))
        lbl_bank_name.grid(row=3, column=1, sticky='e', padx=(pad, pad), pady=pad)
        entry_bank_name = ctk.CTkEntry(bank_win, justify='right', font=font_bank)
        entry_bank_name.grid(row=3, column=0, sticky='ew', padx=(pad, pad), pady=pad)
        entry_bank_name.insert(0, bank_data.get('bank_name', ''))

        # شعبه بانک (Bank Branch)
        lbl_branch = ctk.CTkLabel(bank_win, text='شعبه بانک', font=('vazirmatn', 12, 'bold'))
        lbl_branch.grid(row=4, column=1, sticky='e', padx=(pad, pad), pady=pad)
        entry_branch = ctk.CTkEntry(bank_win, justify='right', font=font_bank)
        entry_branch.grid(row=4, column=0, sticky='ew', padx=(pad, pad), pady=pad)
        entry_branch.insert(0, bank_data.get('bank_branch', ''))

        # شناسه پرداخت (Payment ID)
        lbl_payment_id = ctk.CTkLabel(bank_win, text='شناسه پرداخت', font=('vazirmatn', 12, 'bold'))
        lbl_payment_id.grid(row=5, column=1, sticky='e', padx=(pad, pad), pady=pad)
        entry_payment_id = ctk.CTkEntry(bank_win, justify='right', font=font_bank)
        entry_payment_id.grid(row=5, column=0, sticky='ew', padx=(pad, pad), pady=pad)
        entry_payment_id.insert(0, bank_data.get('payment_id', ''))
        
        # شماره کارت (Card number)
        lbl_card = ctk.CTkLabel(bank_win, text='شماره کارت', font=('vazirmatn', 12, 'bold'))
        lbl_card.grid(row=6, column=1, sticky='e', padx=(pad, pad), pady=pad)
        entry_card = ctk.CTkEntry(bank_win, justify='right', font=font_bank)
        entry_card.grid(row=6, column=0, sticky='ew', padx=(pad, pad), pady=pad)
        entry_card.insert(0, bank_data.get('card_number', ''))

        # Save button
        def save_bank_data():
            bank_data['owner_name'] = entry_owner.get().strip()
            bank_data['account_number'] = entry_account.get().strip()
            bank_data['shaba_number'] = entry_shaba.get().strip()
            bank_data['bank_name'] = entry_bank_name.get().strip()
            bank_data['bank_branch'] = entry_branch.get().strip()
            bank_data['payment_id'] = entry_payment_id.get().strip()
            bank_data['card_number'] = entry_card.get().strip()
            bank_win.destroy()

        btn_save_bank = ctk.CTkButton(bank_win, text='ذخیره', command=save_bank_data, font=('vazirmatn', 12, 'bold'))
        btn_save_bank.grid(row=7, column=0, columnspan=2, pady=(pad, pad), padx=pad, sticky='ew')

    def open_guarantee_window():
        """Open a window for editing contract guarantee details."""
        guarantee_win = ctk.CTkToplevel(top)
        guarantee_win.title('ویرایش تضمین قرارداد')
        guarantee_win.geometry('350x180')
        guarantee_win.resizable(False, False)

        guarantee_win.grid_columnconfigure(1, weight=1)
        font_guarantee = ('vazirmatn', 12, 'bold')
        
        lbl_amount_guarantee = ctk.CTkLabel(guarantee_win, text='مبلغ تضمین', font=font_guarantee)
        lbl_amount_guarantee.grid(row=0, column=3, sticky='e', padx=pad, pady=pad)
        lbl_rial_guarantee = ctk.CTkLabel(guarantee_win, text='ریال', font=font_guarantee)
        lbl_rial_guarantee.grid(row=0, column=0, sticky='w', padx=(4, pad))

        def format_guarantee_amount_edit(event=None):
            text = entry_amount.get().replace(',', '')
            if text.isdigit():
                formatted_text = f"{int(text):,}"
                entry_amount.delete(0, tk.END)
                entry_amount.insert(0, formatted_text)
        
        vcmd_guarantee_edit = (guarantee_win.register(validate_amount_edit), '%P')
        entry_amount = ctk.CTkEntry(guarantee_win, justify='right', font=('vazirmatn', 12), validate='key', validatecommand=vcmd_guarantee_edit)
        entry_amount.grid(row=0, column=1, columnspan=2, sticky='ew', padx=pad, pady=pad)
        entry_amount.insert(0, guarantee_data.get('guarantee_amount', '').replace(',', ''))
        entry_amount.bind("<KeyRelease>", format_guarantee_amount_edit)
        format_guarantee_amount_edit()

        lbl_type = ctk.CTkLabel(guarantee_win, text='نوع ضمانت', font=font_guarantee)
        lbl_type.grid(row=1, column=3, sticky='e', padx=pad, pady=pad)
        guarantee_types = ['چک', 'سفته', 'ضمانت نامه بانکی', 'تعهد', 'سایر']
        combo_type = ctk.CTkComboBox(guarantee_win, values=guarantee_types, font=('vazirmatn', 12), state='readonly')
        combo_type.grid(row=1, column=0, columnspan=3, sticky='ew', padx=pad, pady=pad)
        entry_type_other = ctk.CTkEntry(guarantee_win, justify='right', font=('vazirmatn', 12))
        entry_type_other.grid(row=1, column=0, columnspan=3, sticky='ew', padx=pad, pady=pad)

        def on_guarantee_type_change_edit(choice):
            if choice == 'سایر':
                combo_type.grid_remove()
                entry_type_other.grid()
            else:
                entry_type_other.grid_remove()
                combo_type.grid()

        combo_type.configure(command=on_guarantee_type_change_edit)
        
        gt_val = guarantee_data.get('guarantee_type', '')
        if gt_val in guarantee_types and gt_val != 'سایر':
            combo_type.set(gt_val)
            entry_type_other.grid_remove()
        elif gt_val and gt_val not in guarantee_types:
            # Custom value - show entry field
            combo_type.set('سایر')
            entry_type_other.insert(0, gt_val)
            combo_type.grid_remove()
            entry_type_other.grid()
        else:
            # Empty or 'سایر' - show combobox
            combo_type.set('سایر')
            entry_type_other.grid_remove()

        def save_guarantee_data():
            guarantee_data['guarantee_amount'] = entry_amount.get().strip().replace(',', '')
            guarantee_data['guarantee_type'] = entry_type_other.get().strip() if combo_type.get() == 'سایر' else combo_type.get()
            guarantee_win.destroy()

        btn_save = ctk.CTkButton(guarantee_win, text='ذخیره', command=save_guarantee_data, font=font_guarantee)
        btn_save.grid(row=2, column=0, columnspan=4, sticky='ew', padx=pad, pady=pad)

    # --- Row 3: Contract Amount and Bank Account Button ---
    amount_frame = ctk.CTkFrame(top)
    amount_frame.grid(row=3, column=1, columnspan=2, sticky='ew', padx=pad, pady=pad)
    amount_frame.grid_columnconfigure(0, weight=1)
    amount_frame.grid_columnconfigure(2, weight=1)
    amount_frame.grid_columnconfigure(3, weight=0)

    lbl_contract_amount = ctk.CTkLabel(amount_frame, text='مبلغ قرارداد', font=font_spec)
    lbl_contract_amount.grid(row=0, column=3, sticky='e', padx=(0, pad))
    
    # A sub-frame for the two buttons
    buttons_frame = ctk.CTkFrame(amount_frame, fg_color="transparent")
    buttons_frame.grid(row=0, column=0, sticky='ew', padx=(0, pad))
    buttons_frame.grid_columnconfigure(0, weight=1)
    buttons_frame.grid_columnconfigure(1, weight=1)

    btn_guarantee = ctk.CTkButton(buttons_frame, text='تضمین قرارداد', command=open_guarantee_window, font=font_spec)
    btn_guarantee.grid(row=0, column=0, sticky='ew', padx=(0, 2))

    btn_bank = ctk.CTkButton(buttons_frame, text='ویرایش حساب', command=open_bank_account_window, font=font_spec)
    btn_bank.grid(row=0, column=1, sticky='ew', padx=(2, 0))

    entry_contract_amount = ctk.CTkEntry(amount_frame, justify='right', font=font_entry)
    entry_contract_amount.grid(row=0, column=2, sticky='ew', padx=(0, pad))
    entry_contract_amount.insert(0, data.get('contract_amount') or '')
    
    lbl_rial = ctk.CTkLabel(amount_frame, text='ریال', font=('vazirmatn', 12, 'bold'))
    lbl_rial.grid(row=0, column=1, sticky='w', padx=(0, 4))

    def validate_amount_edit(P):
        return str.isdigit(convert_persian_to_english(P).replace(',', '')) or P == ""

    def format_amount_edit(event=None):
        text = entry_contract_amount.get().replace(',', '')
        if text.isdigit():
            formatted_text = f"{int(text):,}"
            entry_contract_amount.delete(0, tk.END)
            entry_contract_amount.insert(0, formatted_text)

    vcmd_edit = (top.register(validate_amount_edit), '%P')
    entry_contract_amount.configure(validate='key', validatecommand=vcmd_edit)
    entry_contract_amount.bind("<KeyRelease>", format_amount_edit)
    format_amount_edit() # Format initial value

    # --- Row 4: Duration (از - تا) (edit) ---
    # Dedicated row for duration with RTL layout
    duration_frame = ctk.CTkFrame(top)
    duration_frame.grid(row=4, column=1, columnspan=2, sticky='ew', padx=pad, pady=pad)
    duration_frame.grid_columnconfigure(0, weight=1) # calc label
    duration_frame.grid_columnconfigure(1, weight=1) # to entry
    duration_frame.grid_columnconfigure(2, weight=0) # to cal button
    duration_frame.grid_columnconfigure(3, weight=0) # to label
    duration_frame.grid_columnconfigure(4, weight=1) # from entry
    duration_frame.grid_columnconfigure(5, weight=0) # from cal button
    duration_frame.grid_columnconfigure(6, weight=0) # from label

    # RTL layout: from right to left
    lbl_duration_from = ctk.CTkLabel(duration_frame, text='از:', font=font_spec)
    lbl_duration_from.grid(row=0, column=6, sticky='e', padx=(0, 4))

    btn_cal_from = ctk.CTkButton(duration_frame, text='تقویم', command=lambda: open_calendar(top, entry_duration_from), width=60, font=('vazirmatn', 11, 'bold'))
    btn_cal_from.grid(row=0, column=5, sticky='e', padx=(0, 4))

    entry_duration_from = ctk.CTkEntry(duration_frame, justify='right', font=font_entry, width=150)    
    entry_duration_from.grid(row=0, column=4, sticky='ew', padx=(0, pad))
    entry_duration_from.insert(0, data.get('duration_from') or jdatetime.date.today().strftime('%Y-%m-%d'))
    entry_duration_from.bind('<KeyRelease>', normalize_date_input)

    lbl_duration_to = ctk.CTkLabel(duration_frame, text='تا:', font=font_spec)
    lbl_duration_to.grid(row=0, column=3, sticky='e', padx=(0, 4))

    entry_duration_to = ctk.CTkEntry(duration_frame, justify='right', font=font_entry, width=150)
    entry_duration_to.grid(row=0, column=1, sticky='ew', padx=(0, pad))
    entry_duration_to.insert(0, data.get('duration_to') or jdatetime.date.today().strftime('%Y-%m-%d'))
    entry_duration_to.bind('<KeyRelease>', normalize_date_input)

    btn_cal_to = ctk.CTkButton(duration_frame, text='تقویم', command=lambda: open_calendar(top, entry_duration_to), width=60, font=('vazirmatn', 11, 'bold'))
    btn_cal_to.grid(row=0, column=2, sticky='e', padx=(0, 4))

    lbl_duration_calc = ctk.CTkLabel(duration_frame, text='', font=('vazirmatn', 12), text_color='#0066cc')
    lbl_duration_calc.grid(row=0, column=0, sticky='w', padx=(pad, 0))

    def compute_duration_label_edit():
        from_val = entry_duration_from.get().strip()
        to_val = entry_duration_to.get().strip()
        if not from_val or not to_val:
            lbl_duration_calc.configure(text='')
            return
        try:
            d1 = jdatetime.datetime.strptime(from_val, '%Y-%m-%d').date()
            d2 = jdatetime.datetime.strptime(to_val, '%Y-%m-%d').date()
            today = jdatetime.date.today()
            
            # Check if end date is past
            if d2 < today:
                status_var.set('راکد')
            else:
                status_var.set('در جریان')
            
            # convert to Gregorian for timedelta
            g1 = d1.togregorian()
            g2 = d2.togregorian()
            diff_days = (g2 - g1).days
            if diff_days < 0:
                lbl_duration_calc.configure(text='تاریخ پایان قبل از شروع است')
            else:
                # Calculate years, months, days
                years = d2.year - d1.year
                months = d2.month - d1.month
                days = d2.day - d1.day
                if days < 0:
                    months -= 1
                    if d2.month == 1:
                        prev_month_days = 31
                    else:
                        prev_month_days = 31 if d2.month - 1 in [1, 3, 5, 7, 9, 11] else (30 if d2.month - 1 in [4, 6, 8, 10] else 29)
                    days += prev_month_days
                if months < 0:
                    years -= 1
                    months += 12
                lbl_duration_calc.configure(text=f'مدت قرارداد: {years} سال، {months} ماه، {days} روز')
        except Exception:
            lbl_duration_calc.configure(text='فرمت تاریخ باید YYYY-MM-DD باشد')

    entry_duration_from.bind('<FocusOut>', lambda e: compute_duration_label_edit())
    entry_duration_to.bind('<FocusOut>', lambda e: compute_duration_label_edit())
    compute_duration_label_edit()

    # Description
    lbl_desc = ctk.CTkLabel(top, text='توضیحات', font=font_spec)
    lbl_desc.grid(row=5, column=2, sticky='ne', padx=(0, pad), pady=(pad, 0))
    text_desc = ctk.CTkTextbox(top, height=150, font=font_entry)
    # text_desc.tag_configure("right", justify="right")
    text_desc.grid(row=5, column=1, sticky='nsew', padx=pad, pady=(0, pad))
    text_desc.insert('1.0', data.get('description') or '')
    top.grid_rowconfigure(5, weight=1)

    # Attachments + Export Checkbox + Status
    # Attachments + Export Checkbox + Status
    files_frame = ctk.CTkFrame(top)
    files_frame.grid(row=6, column=0, columnspan=3, sticky='ew', padx=pad, pady=pad)
    files_frame.grid_columnconfigure(0, weight=1)  # File label expands to left
    files_frame.grid_columnconfigure(1, weight=0)  # Attach button
    files_frame.grid_columnconfigure(2, weight=0)  # Export checkbox
    files_frame.grid_columnconfigure(3, weight=0)  # Status radios
    files_frame.grid_columnconfigure(4, weight=0)  # Button stays small on right

    lbl_files = ctk.CTkLabel(files_frame, text='فایلی انتخاب نشده است', font=('vazirmatn', 12), text_color="gray")
    lbl_files.grid(row=0, column=0, sticky='w', padx=(pad, 0))

    def select_files():
        files = filedialog.askopenfilenames(title='انتخاب فایل‌ها')
        if files:
            selected_files['list'] = list(files)
            lbl_files.configure(text=f'{len(selected_files["list"])} فایل انتخاب شد')

    btn_attach = ctk.CTkButton(files_frame, text='افزودن فایل پیوست', command=select_files, font=('vazirmatn', 13, 'bold'))
    btn_attach.grid(row=0, column=1, sticky='e', padx=(pad, 0))

    # Export checkbox (default ON)
    export_var = tk.BooleanVar(value=True) # Default to ON for edit mode
    chk_export = ctk.CTkCheckBox(
        files_frame,
        text='بروزرسانی فایل XLSX',
        variable=export_var,
        font=('vazirmatn', 11)
    )
    chk_export.grid(row=0, column=2, sticky='e', padx=pad)

    # Status indicator (Radios)
    status_var = tk.StringVar(value=data.get('status', 'در جریان'))
    
    frame_status = ctk.CTkFrame(files_frame, fg_color="transparent")
    frame_status.grid(row=0, column=3, sticky='e', padx=pad)
    
    lbl_status = ctk.CTkLabel(frame_status, text='وضعیت:', font=('vazirmatn', 11, 'bold'))
    lbl_status.pack(side='right', padx=(0, 8))
    
    rb_active = ctk.CTkRadioButton(frame_status, text='در جریان', variable=status_var, value='در جریان', font=('vazirmatn', 11), text_color='green')
    rb_active.pack(side='right', padx=2)
    
    rb_inactive = ctk.CTkRadioButton(frame_status, text='راکد', variable=status_var, value='راکد', font=('vazirmatn', 11), text_color='red')
    rb_inactive.pack(side='right', padx=2)

    # Save button for edit
    btn_save = ctk.CTkButton(top, text='ذخیره تغییرات', command=lambda: save_edit_case(top, case_id, selected_files, entry_title, entry_date, entry_duration_from, entry_duration_to, entry_mojer, entry_mostajjer, entry_karfarma, entry_piman, entry_subject, text_desc, entry_contract_amount, combo_case_type, entry_case_type_other, bank_data, status_var, export_var, guarantee_data), font=('vazirmatn', 14, 'bold'), height=40)
    btn_save.grid(row=7, column=0, columnspan=3, sticky='ew', padx=pad, pady=(pad, pad))


def save_edit_case(top, case_id, selected_files, entry_title, entry_date, entry_duration_from, entry_duration_to, entry_mojer, entry_mostajjer, entry_karfarma, entry_piman, entry_subject, text_desc, entry_contract_amount, combo_case_type, entry_case_type_other, bank_data, status_var, export_var=None, guarantee_data=None):
    title = entry_title.get().strip()
    if not title:
        messagebox.showwarning('خطا', 'عنوان پرونده الزامی است', parent=top)
        return

    # fetch existing record to get folder
    cur = get_case_by_id(case_id)
    if not cur:
        messagebox.showerror('خطا', 'پرونده پیدا نشد', parent=top)
        return

    folder = cur.get('folder_path')
    if not folder:
        # ensure we have a folder (create if missing)
        folder = os.path.join(UPLOADS_DIR, case_id)
        os.makedirs(folder, exist_ok=True)

    # copy new attachments (if any)
    for f in selected_files['list']:
        try:
            shutil.copy2(f, folder)
        except Exception:
            pass

    # compute duration string
    from_val = entry_duration_from.get().strip()
    to_val = entry_duration_to.get().strip()
    duration_str = ''
    try:
        if from_val and to_val:
            d1 = jdatetime.datetime.strptime(from_val, '%Y-%m-%d').date()
            d2 = jdatetime.datetime.strptime(to_val, '%Y-%m-%d').date()
            diff_days = (d2.togregorian() - d1.togregorian()).days
            if diff_days >= 0:
                duration_str = f'{convert_english_to_persian(str(diff_days))} \u200eروز'
            else:
                duration_str = ''
    except Exception:
        duration_str = ''

    case_type_val = combo_case_type.get()
    if case_type_val == 'سایر':
        case_type_val = entry_case_type_other.get().strip()

    data = {
        'title': title,
        'date': entry_date.get().strip() or cur.get('date'),
        'duration': duration_str,
        'duration_from': from_val,
        'duration_to': to_val,
        'mojer': entry_mojer.get().strip(),
        'mostajjer': entry_mostajjer.get().strip(),
        'karfarma': entry_karfarma.get().strip(),
        'piman': entry_piman.get().strip(),
        'subject': entry_subject.get().strip(),
        'contract_amount': entry_contract_amount.get().strip().replace(',', ''),
        'bank_owner_name': bank_data.get('owner_name', '').strip(),
        'bank_account_number': bank_data.get('account_number', '').strip(),
        'bank_shaba_number': bank_data.get('shaba_number', '').strip(),
        'bank_card_number': bank_data.get('card_number', '').strip(),
        'bank_name': bank_data.get('bank_name', '').strip(),
        'bank_branch': bank_data.get('bank_branch', '').strip(),
        'payment_id': bank_data.get('payment_id', '').strip(),        
        'guarantee_amount': guarantee_data.get('guarantee_amount', '').strip() if guarantee_data else '',
        'guarantee_type': guarantee_data.get('guarantee_type', '').strip() if guarantee_data else '',
        'description': text_desc.get('1.0', tk.END).strip(),
        'folder_path': folder,
        'case_type': case_type_val,
        'status': status_var.get()
    }

    # Export to CSV and TXT if checkbox is enabled
    if export_var and export_var.get():
        data['id'] = case_id
        export_case_to_files(data)

    try:
        update_case(case_id, data)
        messagebox.showinfo('موفق', 'ویرایش ذخیره شد', parent=top)
        top.master.deiconify()
        top.destroy()
    except Exception as e:
        messagebox.showerror('خطا', f'خطا در ذخیره: {e}', parent=top)
