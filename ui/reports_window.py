import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import jdatetime
import os
import csv
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import arabic_reshaper
from bidi.algorithm import get_display

import tempfile
from database import get_connection


def open_reports_window(master):
    """Open the Reports window for filtering cases by date range and exporting to PDF."""
    top = ctk.CTkToplevel(master)
    top.title('📊 گزارش گیری — سیستم مدیریت پرونده مهاجر')
    top.geometry('1150x600')

    # --- Simple Grid Layout (RTL) ---
    top.grid_columnconfigure(0, weight=1)  # Flexible space (right side in RTL)
    top.grid_rowconfigure(1, weight=1)    # Treeview expandable

    pad = 8
    font_bold = ('vazirmatn', 12, 'bold')
    font_normal = ('vazirmatn', 12)

    # --- Header Frame ---
    header_frame = ctk.CTkFrame(top)
    header_frame.grid(row=0, column=0, sticky='ew', padx=pad, pady=pad)
    header_frame.grid_columnconfigure(3, weight=1)  # Flexible space in middle
    
    # Buttons (left side - columns 0-2)
    btn_export_xlsx = ctk.CTkButton(header_frame, text='📄 خروجی XLSX', font=font_bold, state=tk.DISABLED, command=lambda: export_to_xlsx(current_data['rows']))
    btn_export_xlsx.grid(row=0, column=1, padx=(0, 4))

    btn_filter = ctk.CTkButton(header_frame, text='🔍 فیلتر', font=font_bold)
    btn_filter.grid(row=0, column=2, padx=(0, pad))

    # Spacer (flexible) - column 3

    # Date inputs (right side - columns 4-7)
    lbl_to = ctk.CTkLabel(header_frame, text='تا تاریخ', font=font_bold)
    lbl_to.grid(row=0, column=4, padx=(pad, 0))

    entry_to = ctk.CTkEntry(header_frame, width=150, justify='right', font=font_normal)
    entry_to.insert(0, jdatetime.date.today().strftime('%Y-%m-%d'))
    entry_to.grid(row=0, column=5, padx=(0, pad))

    lbl_from = ctk.CTkLabel(header_frame, text='از تاریخ', font=font_bold)
    lbl_from.grid(row=0, column=6, padx=(pad, 0))

    entry_from = ctk.CTkEntry(header_frame, width=150, justify='right', font=font_normal)
    entry_from.insert(0, jdatetime.date.today().strftime('%Y-%m-%d'))
    entry_from.grid(row=0, column=7, padx=(0, pad))

    # --- Results Treeview ---
    cols = ('id', 'title', 'subject', 'date', 'case_type', 'duration', 'contract_amount')
    tree = ttk.Treeview(top, columns=cols, show='headings', height=20)

    tree.heading('id', text='شناسه بایگانی')
    tree.heading('title', text='عنوان')
    tree.heading('subject', text='موضوع')
    tree.heading('date', text='تاریخ')
    tree.heading('case_type', text='نوع پرونده')
    tree.heading('duration', text='مدت')
    tree.heading('contract_amount', text='مبلغ قرارداد')

    # Bind heading clicks for sorting
    for col in cols:
        tree.heading(col, command=lambda c=col: sort_tree(c))

    tree.column('id', width=100)
    tree.column('title', width=130)
    tree.column('subject', width=130)
    tree.column('date', width=100)
    tree.column('case_type', width=110)
    tree.column('duration', width=90)
    tree.column('contract_amount', width=110)

    tree.grid(row=1, column=0, padx=pad, pady=pad, sticky='nsew')

    # Scrollbars
    v_scroll = ctk.CTkScrollbar(top, orientation='vertical', command=tree.yview)
    v_scroll.grid(row=1, column=1, sticky='ns', padx=(0, pad), pady=pad)
    h_scroll = ctk.CTkScrollbar(top, orientation='horizontal', command=tree.xview)
    h_scroll.grid(row=2, column=0, sticky='ew', padx=pad, pady=(0, pad))
    tree.configure(yscroll=v_scroll.set, xscroll=h_scroll.set)

    # Status label
    lbl_status = ctk.CTkLabel(top, text='لطفا فیلتر را اعمال کنید', font=('vazirmatn', 11), text_color='gray')
    lbl_status.grid(row=3, column=0, padx=pad, pady=(pad, pad), sticky='w')

    # --- Sort Controls Frame ---
    sort_frame = ctk.CTkFrame(top, fg_color="transparent")
    sort_frame.grid(row=3, column=0, columnspan=1, padx=pad, pady=(pad, pad), sticky='e')
    
    # Sort by label
    lbl_sort_by = ctk.CTkLabel(sort_frame, text='مرتب‌سازی بر اساس:', font=('vazirmatn', 11, 'bold'))
    lbl_sort_by.pack(side='right', padx=(0, 8))
    
    # Sort column selection (radio buttons)
    sort_col_var = tk.StringVar(value='id')
    sort_cols_frame = ctk.CTkFrame(sort_frame, fg_color="transparent")
    sort_cols_frame.pack(side='right', padx=(0, 12))
    
    col_labels = {
        'id': 'شناسه بایگانی',
        'title': 'عنوان',
        'subject': 'موضوع',
        'date': 'تاریخ',
        'case_type': 'نوع پرونده',
        'duration': 'مدت',
        'contract_amount': 'مبلغ قرارداد'
    }
    
    for col in cols:
        rb = ctk.CTkRadioButton(sort_cols_frame, text=col_labels[col], variable=sort_col_var, value=col,
                           font=('vazirmatn', 10), command=lambda: apply_sort_controls())
        rb.pack(side='right', padx=2)
    
    # Sort direction separator
    sep = ctk.CTkLabel(sort_frame, text='|', font=('vazirmatn', 11))
    sep.pack(side='right', padx=8)
    
    # Sort direction label
    lbl_sort_dir = ctk.CTkLabel(sort_frame, text='نوع مرتب‌سازی:', font=('vazirmatn', 11, 'bold'))
    lbl_sort_dir.pack(side='right', padx=(0, 8))
    
    # Sort direction selection (radio buttons)
    sort_dir_var = tk.StringVar(value='asc')
    
    rb_asc = ctk.CTkRadioButton(sort_frame, text='کوچک به بزرگ ▲', variable=sort_dir_var, value='asc',
                           font=('vazirmatn', 10), command=lambda: apply_sort_controls())
    rb_asc.pack(side='right', padx=2)
    
    rb_desc = ctk.CTkRadioButton(sort_frame, text='بزرگ به کوچک ▼', variable=sort_dir_var, value='desc',
                            font=('vazirmatn', 10), command=lambda: apply_sort_controls())
    rb_desc.pack(side='right', padx=2)

    current_data = {'rows': [], 'sort_column': None, 'sort_reverse': False}

    def apply_sort_controls():
        """Apply sorting based on the radio button selections."""
        col = sort_col_var.get()
        reverse = sort_dir_var.get() == 'desc'
        
        items = [(tree.set(k, col), k) for k in tree.get_children('')]
        
        # Try to sort as numbers first, fallback to strings
        try:
            items.sort(key=lambda x: float(x[0]) if x[0] != '-----' else float('-inf'), reverse=reverse)
        except ValueError:
            items.sort(key=lambda x: x[0], reverse=reverse)
        
        # Rearrange items in sorted position
        for index, (val, k) in enumerate(items):
            tree.move(k, '', index)

    def sort_tree(col):
        """Sort the treeview by column, toggling between ascending and descending."""
        # Get all items
        items = [(tree.set(k, col), k) for k in tree.get_children('')]
        
        # Check if we're sorting the same column (toggle direction)
        if current_data['sort_column'] == col:
            current_data['sort_reverse'] = not current_data['sort_reverse']
        else:
            current_data['sort_column'] = col
            current_data['sort_reverse'] = False
        
        # Try to sort as numbers first, fallback to strings
        try:
            items.sort(key=lambda x: float(x[0]) if x[0] != '-----' else float('-inf'), reverse=current_data['sort_reverse'])
        except ValueError:
            items.sort(key=lambda x: x[0], reverse=current_data['sort_reverse'])
        
        # Rearrange items in sorted position
        for index, (val, k) in enumerate(items):
            tree.move(k, '', index)
        
        # Update column header to show sort direction
        arrow = ' ▼' if current_data['sort_reverse'] else ' ▲'
        for c in cols:
            heading_text = tree.heading(c, 'text').replace(' ▼', '').replace(' ▲', '')
            if c == col:
                tree.heading(c, text=heading_text + arrow)
            else:
                tree.heading(c, text=heading_text)

    def do_filter():
        """Filter cases by date range."""
        date_from = entry_from.get().strip()
        date_to = entry_to.get().strip()

        if not date_from or not date_to:
            messagebox.showwarning('هشدار', 'لطفا هر دو تاریخ را وارد کنید', parent=top)
            return

        try:
            # Validate dates
            d_from = jdatetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            d_to = jdatetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            if d_from > d_to:
                messagebox.showwarning('هشدار', 'تاریخ شروع باید قبل از تاریخ پایان باشد', parent=top)
                return
        except Exception as e:
            messagebox.showerror('خطا', f'فرمت تاریخ نامعتبر: {e}', parent=top)
            return

        # Query database for cases in date range
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('''SELECT id, title, subject, date, case_type, duration, contract_amount 
                       FROM cases 
                       WHERE date >= ? AND date <= ? 
                       ORDER BY date DESC''', (date_from, date_to))
        rows = cur.fetchall()
        conn.close()

        # Clear previous results
        for i in tree.get_children():
            tree.delete(i)

        # Insert new results
        for r in rows:
            vals = [str(r[i]) if r[i] else '-----' for i in range(7)]
            tree.insert('', 'end', values=vals)

        current_data['rows'] = rows
        current_data['sort_column'] = None  # Reset sort state
        current_data['sort_reverse'] = False
        lbl_status.configure(text=f'تعداد پرونده‌های یافت شده: {len(rows)}', text_color='green')
        
        # Enable export buttons if there are results
        btn_export_xlsx.configure(state='normal' if rows else 'disabled')

    btn_filter.configure(command=do_filter)

    def export_to_xlsx(rows_to_export):
        """Export current table to an XLSX file."""
        if not rows_to_export:
            messagebox.showwarning('هشدار', 'جدولی برای خروجی موجود نیست', parent=top)
            return

        file_path = filedialog.asksaveasfilename(
            parent=top,
            defaultextension='.xlsx',
            filetypes=[('Excel Files', '*.xlsx'), ('All Files', '*.*')],
            initialfile=f"گزارش_پرونده‌ها_{jdatetime.date.today().strftime('%Y%m%d')}.xlsx"
        )

        if not file_path:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "گزارش پرونده ها"
            sheet.sheet_view.rightToLeft = True

            # --- Styles ---
            bold_font = Font(name='Vazirmatn', bold=True)
            regular_font = Font(name='Vazirmatn')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_alignment = Alignment(horizontal='center', vertical='center')

            headers = ['شناسه بایگانی', 'عنوان', 'موضوع', 'تاریخ', 'نوع پرونده', 'مدت', 'مبلغ قرارداد']
            headers.insert(0, 'ردیف') # Add Row column header
            reshaped_headers = [get_display(arabic_reshaper.reshape(h)) for h in headers]
            sheet.append(reshaped_headers)

            # Apply style to header
            for col_num, header in enumerate(reshaped_headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = bold_font
                cell.border = thin_border
                cell.alignment = center_alignment

            # Apply style to data rows
            for row_idx, row in enumerate(rows_to_export, start=2):
                # Add row number to the start of the row data
                row_with_num = [row_idx - 1] + list(row)
                reshaped_row = [get_display(arabic_reshaper.reshape(str(item) if item is not None else '')) for item in row_with_num]
                sheet.append(reshaped_row)
                for col_idx, _ in enumerate(reshaped_row, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.font = regular_font
                    cell.border = thin_border
                    cell.alignment = center_alignment

            # Adjust column widths
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width

            workbook.save(file_path)
            messagebox.showinfo('موفق', f'گزارش با موفقیت صادر شد:\n{file_path}', parent=top)
        except Exception as e:
            messagebox.showerror('خطا', f'خطا در صادر کردن فایل XLSX: {str(e)}', parent=top)
