import os
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import jdatetime
import csv
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from database import search_cases, delete_case, get_connection
from ui.details_window import open_details_window

PERSIAN_TO_ENGLISH_MAP = {
    '۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4',
    '۵': '5', '۶': '6', '۷': '7', '۸': '8', '۹': '9'
}

def convert_persian_to_english(text):
    """Converts Persian/Arabic numerals in a string to English numerals."""
    for p, e in PERSIAN_TO_ENGLISH_MAP.items():
        text = text.replace(p, e)
    return text

def convert_english_to_persian(text):
    """Converts English numerals in a string to Persian numerals."""
    english_to_persian_map = {v: k for k, v in PERSIAN_TO_ENGLISH_MAP.items()}
    for e, p in english_to_persian_map.items():
        text = text.replace(e, p)
    return text

def calculate_duration_text(duration_from, duration_to):
    """
    Calculate duration in 'X سال، Y ماه، Z روز' format from two dates.
    Returns the formatted string or '-----' if dates are invalid.
    """
    if not duration_from or not duration_to:
        return '-----'
    
    try:
        d1 = jdatetime.datetime.strptime(str(duration_from), '%Y-%m-%d').date()
        d2 = jdatetime.datetime.strptime(str(duration_to), '%Y-%m-%d').date()
        
        # Convert to Gregorian for timedelta check
        g1 = d1.togregorian()
        g2 = d2.togregorian()
        diff_days = (g2 - g1).days
        
        if diff_days < 0:
            return 'تاریخ پایان قبل از شروع'
        
        # Add LRM (\u200e) to fix display in Excel
        return f'{convert_english_to_persian(str(diff_days))} \u200eروز'
    except Exception as e:
        return '-----'

class CustomJalaliCalendar(ctk.CTkFrame):
    """A custom Jalali calendar widget built with customtkinter."""
    def __init__(self, master, on_select=None, year=None, month=None, day=None):
        super().__init__(master)
        self.on_select = on_select

        today = jdatetime.date.today()
        self.year = year if year else today.year
        self.month = month if month else today.month
        self.day = day if day else today.day

        self.month_names = [
            "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
            "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
        ]
        self.weekday_names = ["ش", "ی", "د", "س", "چ", "پ", "ج"]

        # --- Header: Month/Year selection ---
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(pady=5, fill='x')
        header_frame.grid_columnconfigure(1, weight=1)
        header_frame.grid_columnconfigure(2, weight=1)

        btn_prev = ctk.CTkButton(header_frame, text=">", width=30, command=self.prev_month)
        btn_prev.grid(row=0, column=0, padx=5)

        self.year_combo = ctk.CTkComboBox(header_frame, values=[str(y) for y in range(today.year - 10, today.year + 10)], command=self.update_calendar)
        self.year_combo.set(str(self.year))
        self.year_combo.grid(row=0, column=1, padx=5, sticky='ew')

        self.month_combo = ctk.CTkComboBox(header_frame, values=self.month_names, command=self.update_calendar)
        self.month_combo.set(self.month_names[self.month - 1])
        self.month_combo.grid(row=0, column=2, padx=5, sticky='ew')

        btn_next = ctk.CTkButton(header_frame, text="<", width=30, command=self.next_month)
        btn_next.grid(row=0, column=3, padx=5)

        # --- Calendar Body ---
        self.calendar_frame = ctk.CTkFrame(self)
        self.calendar_frame.pack(expand=True, fill="both", padx=5, pady=5)

        self.create_calendar()

    def create_calendar(self):
        # Clear previous widgets
        for widget in self.calendar_frame.winfo_children():
            widget.destroy()

        # Weekday headers
        for i, name in enumerate(self.weekday_names):
            self.calendar_frame.grid_columnconfigure(i, weight=1)
            lbl = ctk.CTkLabel(self.calendar_frame, text=name, font=('vazirmatn', 12, 'bold'))
            lbl.grid(row=0, column=i, sticky='nsew')

        # Get calendar data
        first_day_of_month = jdatetime.date(self.year, self.month, 1)
        start_weekday = first_day_of_month.weekday()  # jdatetime: 0=Saturday, 6=Friday. Matches our grid.
        days_in_month = jdatetime.j_days_in_month[self.month - 1]
        if self.month == 12 and not jdatetime.date(self.year, 1, 1).isleap():
            days_in_month = 29

        day_num = 1
        for r in range(1, 7): # Max 6 rows
            self.calendar_frame.grid_rowconfigure(r, weight=1)
            for c in range(7):
                if (r == 1 and c < start_weekday) or day_num > days_in_month:
                    continue
                
                btn = ctk.CTkButton(
                    self.calendar_frame,
                    text=str(day_num),
                    command=lambda d=day_num: self.select_date(d)
                )
                btn.grid(row=r, column=c, sticky='nsew', padx=1, pady=1)
                day_num += 1

    def update_calendar(self, event=None):
        self.year = int(self.year_combo.get())
        self.month = self.month_names.index(self.month_combo.get()) + 1
        self.create_calendar()

    def next_month(self):
        self.month += 1
        if self.month > 12:
            self.month = 1
            self.year += 1
        self.year_combo.set(str(self.year))
        self.month_combo.set(self.month_names[self.month - 1])
        self.update_calendar()

    def prev_month(self):
        self.month -= 1
        if self.month < 1:
            self.month = 12
            self.year -= 1
        self.year_combo.set(str(self.year))
        self.month_combo.set(self.month_names[self.month - 1])
        self.update_calendar()

    def select_date(self, day):
        if self.on_select:
            date_str = f"{self.year}-{self.month:02d}-{day:02d}"
            self.on_select(date_str)

def open_calendar(master, entry_widget):
    """Opens a Jalali calendar to select a date and inserts it into the entry_widget."""
    def on_date_select(date_str):
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, date_str)
        cal_win.destroy()
        entry_widget.event_generate('<FocusOut>')

    cal_win = ctk.CTkToplevel(master)
    cal_win.title("انتخاب تاریخ")
    cal_win.geometry("400x350")
    cal_win.resizable(False, False)

    try:
        current_date_str = entry_widget.get()
        year, month, day = map(int, current_date_str.split('-'))
        cal = CustomJalaliCalendar(cal_win, on_select=on_date_select, year=year, month=month, day=day)
    except (ValueError, IndexError):
        cal = CustomJalaliCalendar(cal_win, on_select=on_date_select)

    cal.pack(pady=10, padx=10, fill="both", expand=True)


def open_search_records(master):
    """Open the Search Records window as a Toplevel and provide search UI."""
    master.withdraw()
    top = ctk.CTkToplevel(master)
    top.title('جستجوی پرونده‌ها — سیستم مدیریت پرونده مهاجر')
    top.geometry('1460x600')

    def on_close():
        master.deiconify()
        top.destroy()
    
    top.protocol("WM_DELETE_WINDOW", on_close)

    # --- RTL Layout Configuration ---
    top.grid_columnconfigure(0, weight=1)  # Main content expands
    top.grid_rowconfigure(1, weight=1)    # Treeview expands

    pad = 8

    # --- Top Control Bar (Frame for all controls) ---
    control_frame = ctk.CTkFrame(top)
    control_frame.grid(row=0, column=0, sticky='ew', padx=pad, pady=pad)
    control_frame.grid_columnconfigure(4, weight=1)  # Spacer expands

    # Left side buttons (right-to-left: delete, text, csv)
    btn_delete = ctk.CTkButton(control_frame, text='حذف پرونده', command=lambda: delete_selected_records(), font=('vazirmatn', 11, 'bold'), state="disabled")
    btn_delete.grid(row=0, column=0, padx=(0, pad))

    btn_export_xlsx = ctk.CTkButton(control_frame, text='خروجی فایل XLSX', font=('vazirmatn', 11, 'bold'), state="disabled")
    btn_export_xlsx.grid(row=0, column=1, padx=(0, pad))

    # Search button
    btn_search = ctk.CTkButton(control_frame, text='جستجوی پرونده ها', command=lambda: do_search(), font=('vazirmatn', 11, 'bold'))
    btn_search.grid(row=0, column=3, padx=(0, pad))

    # Search entry (column 5 - spans date area)
    entry_q = ctk.CTkEntry(control_frame, width=750, justify='right', font=('vazirmatn', 12))
    entry_q.grid(row=0, column=5, columnspan=8, padx=(pad, 2), sticky='ew')

    def on_filter_change(choice):
        """Show/hide date entries based on filter selection."""
        selected_filter = choice
        if selected_filter == 'بر اساس تاریخ':
            # Show date entries, hide text entry
            entry_q.grid_remove()
            
            # RTL Layout: [Entry][Today][Cal][Label]
            
            # To Date Block (Left) - Cols 5, 6, 7, 8
            entry_date_to.grid(row=0, column=5, padx=(2, 2), sticky='ew')
            btn_today_to.grid(row=0, column=6, padx=(2, 2), sticky='ew')
            btn_cal_to.grid(row=0, column=7, padx=(2, 2), sticky='ew')
            lbl_date_to.grid(row=0, column=8, padx=(2, 15), sticky='e') # Extra pad to separate blocks

            # From Date Block (Right) - Cols 9, 10, 11, 12
            entry_date_from.grid(row=0, column=9, padx=(2, 2), sticky='ew')
            btn_today_from.grid(row=0, column=10, padx=(2, 2), sticky='ew')
            btn_cal_from.grid(row=0, column=11, padx=(2, 2), sticky='ew')
            lbl_date_from.grid(row=0, column=12, padx=(2, 0), sticky='e')
            
        else:
            # Show text entry, hide date entries
            entry_q.grid()
            lbl_date_from.grid_remove()
            entry_date_from.grid_remove()
            btn_cal_from.grid_remove()
            btn_today_from.grid_remove()

            lbl_date_to.grid_remove()
            entry_date_to.grid_remove()
            btn_cal_to.grid_remove()
            btn_today_to.grid_remove()

    # Combobox filter
    combo = ctk.CTkComboBox(control_frame, values=['جستجوی کلی', 'بر اساس عنوان', 'بر اساس موضوع', 'بر اساس تاریخ', 'بر اساس نوع پرونده', 'بر اساس شناسه بایگانی'], font=('vazirmatn', 12, 'bold'), width=150, state='readonly', command=on_filter_change)
    combo.set('جستجوی کلی')  # Set to "جستجوی کلی" as default
    combo.grid(row=0, column=13, padx=(pad, 0))

    # Filter label
    lbl_filter = ctk.CTkLabel(control_frame, text='نوع فیلتر', font=('vazirmatn', 12, 'bold'))
    lbl_filter.grid(row=0, column=14, padx=(pad, 0))

    def normalize_date_input(event):
        widget = event.widget
        current_text = widget.get()
        new_text = convert_persian_to_english(current_text)
        if current_text != new_text:
            widget.delete(0, tk.END)
            widget.insert(0, new_text)

    # --- Date range entries (hidden by default) ---
    # From Date controls
    entry_date_from = ctk.CTkEntry(control_frame, width=120, justify='right', font=('vazirmatn', 12))
    entry_date_from.insert(0, jdatetime.date.today().strftime('%Y-%m-%d'))
    entry_date_from.bind('<KeyRelease>', normalize_date_input)
    
    btn_today_from = ctk.CTkButton(control_frame, text='امروز', command=lambda: (entry_date_from.delete(0, tk.END), entry_date_from.insert(0, jdatetime.date.today().strftime('%Y-%m-%d'))), font=('vazirmatn', 11, 'bold'), width=50)
    btn_cal_from = ctk.CTkButton(control_frame, text='تقویم', command=lambda: open_calendar(top, entry_date_from), width=40, font=('vazirmatn', 11, 'bold'))
    
    # To Date controls
    entry_date_to = ctk.CTkEntry(control_frame, width=120, justify='right', font=('vazirmatn', 12))
    entry_date_to.insert(0, jdatetime.date.today().strftime('%Y-%m-%d'))
    entry_date_to.bind('<KeyRelease>', normalize_date_input)
    
    btn_today_to = ctk.CTkButton(control_frame, text='امروز', command=lambda: (entry_date_to.delete(0, tk.END), entry_date_to.insert(0, jdatetime.date.today().strftime('%Y-%m-%d'))), font=('vazirmatn', 11, 'bold'), width=50)
    btn_cal_to = ctk.CTkButton(control_frame, text='تقویم', command=lambda: open_calendar(top, entry_date_to), width=40, font=('vazirmatn', 11, 'bold'))
    
    lbl_date_from = ctk.CTkLabel(control_frame, text='از تاریخ', font=('vazirmatn', 11, 'bold'))
    lbl_date_to = ctk.CTkLabel(control_frame, text='تا تاریخ', font=('vazirmatn', 11, 'bold'))

    # on_filter_change is defined above and passed to command=


    # --- RTL Treeview ---
    # Columns order must match database query: id, title, subject, date, case_type, duration, status, contract_amount
    cols = ('id', 'title', 'subject', 'date', 'case_type', 'duration', 'status', 'contract_amount')
    tree = ttk.Treeview(top, columns=cols, show='headings')
    # Headings
    tree.heading('title', text='عنوان')
    tree.heading('subject', text='موضوع')
    tree.heading('case_type', text='نوع پرونده')
    tree.heading('date', text='تاریخ')
    tree.heading('duration', text='مدت')
    tree.heading('status', text='وضعیت')
    tree.heading('contract_amount', text='مبلغ قرارداد')
    tree.heading('id', text='شناسه بایگانی')

    # Set column widths for all columns
    tree.column('id', width=140, minwidth=100)
    tree.column('title', width=120, minwidth=100)
    tree.column('subject', width=120, minwidth=100)
    tree.column('date', width=100, minwidth=90)
    tree.column('case_type', width=100, minwidth=90)
    tree.column('duration', width=80, minwidth=70)
    tree.column('status', width=90, minwidth=80)
    tree.column('contract_amount', width=120, minwidth=100)
    tree.grid(row=1, column=0, padx=pad, pady=pad, sticky='nsew')

    # Bind heading clicks for sorting
    for col in cols:
        tree.heading(col, command=lambda c=col: sort_tree(c))

    # Add scrollbars for horizontal and vertical scrolling
    vsb = ttk.Scrollbar(top, orient='vertical', command=tree.yview)
    hsb = ttk.Scrollbar(top, orient='horizontal', command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    
    vsb.grid(row=1, column=1, sticky='ns', padx=(0, pad), pady=pad)
    hsb.grid(row=2, column=0, sticky='ew', padx=pad, pady=(0, pad))

    # Status label for showing number of results
    lbl_status = ctk.CTkLabel(top, text='لطفا جستجو کنید', font=('vazirmatn', 11), text_color='gray')
    lbl_status.grid(row=3, column=0, padx=pad, pady=(pad, pad), sticky='w')

    # --- Sort Controls Frame ---
    sort_frame = ctk.CTkFrame(top)
    sort_frame.grid(row=3, column=0, padx=pad, pady=(pad, pad), sticky='e')
    
    # Sort by label
    lbl_sort_by = ctk.CTkLabel(sort_frame, text='مرتب‌سازی بر اساس:', font=('vazirmatn', 11, 'bold'))
    lbl_sort_by.pack(side='right', padx=(0, 8))
    
    # Sort column selection (radio buttons)
    sort_col_var = tk.StringVar(value='id')
    sort_cols_frame = ctk.CTkFrame(sort_frame, fg_color="transparent")
    sort_cols_frame.pack(side='right', padx=(0, 12))
    
    col_labels = {
        'id': 'شناسه بایگانی',
        'title': 'عنوان',
        'subject': 'موضوع',
        'date': 'تاریخ',
        'case_type': 'نوع پرونده',
        'duration': 'مدت',
        'status': 'وضعیت',
        'contract_amount': 'مبلغ قرارداد'
    }
    
    for col in cols:
        rb = ctk.CTkRadioButton(sort_cols_frame, text=col_labels[col], variable=sort_col_var, value=col, 
                           font=('vazirmatn', 10), command=lambda: apply_sort_controls())
        rb.pack(side='right', padx=2)
    
    # Sort direction separator
    sep = ctk.CTkLabel(sort_frame, text='|', font=('vazirmatn', 11))
    sep.pack(side='right', padx=8)
    
    # Sort direction label
    lbl_sort_dir = ctk.CTkLabel(sort_frame, text='نوع مرتب‌سازی:', font=('vazirmatn', 11, 'bold'))
    lbl_sort_dir.pack(side='right', padx=(0, 8))
    
    # Sort direction selection (radio buttons)
    sort_dir_var = tk.StringVar(value='asc')
    
    rb_asc = ctk.CTkRadioButton(sort_frame, text='کوچک به بزرگ ▲', variable=sort_dir_var, value='asc',
                           font=('vazirmatn', 10), command=lambda: apply_sort_controls())
    rb_asc.pack(side='right', padx=2)
    
    rb_desc = ctk.CTkRadioButton(sort_frame, text='بزرگ به کوچک ▼', variable=sort_dir_var, value='desc',
                            font=('vazirmatn', 10), command=lambda: apply_sort_controls())
    rb_desc.pack(side='right', padx=2)

    # Sort state and current data
    sort_state = {'column': None, 'reverse': False}
    current_data = {'rows': []}

    def update_sort_ui():
        """Update radio buttons and column headers to match sort_state."""
        # Update radio buttons to match sort_state
        sort_col_var.set(sort_state['column'] if sort_state['column'] else 'id')
        sort_dir_var.set('desc' if sort_state['reverse'] else 'asc')
        
        # Update column header arrows
        arrow = ' ▼' if sort_state['reverse'] else ' ▲'
        for c in cols:
            heading_text = tree.heading(c, 'text').replace(' ▼', '').replace(' ▲', '')
            if c == sort_state['column']:
                tree.heading(c, text=heading_text + arrow)
            else:
                tree.heading(c, text=heading_text)
    
    def sort_tree(col):
        """Sort the treeview by column, toggling between ascending and descending."""
        # Get all items
        items = [(tree.set(k, col), k) for k in tree.get_children('')]
        
        # Check if we're sorting the same column (toggle direction)
        if sort_state['column'] == col:
            sort_state['reverse'] = not sort_state['reverse']
        else:
            sort_state['column'] = col
            sort_state['reverse'] = False
        
        # Try to sort as numbers first, fallback to strings
        try:
            items.sort(key=lambda x: float(x[0]) if x[0] != '-----' else float('-inf'), reverse=sort_state['reverse'])
        except ValueError:
            items.sort(key=lambda x: x[0], reverse=sort_state['reverse'])
        
        # Rearrange items in sorted position
        for index, (val, k) in enumerate(items):
            tree.move(k, '', index)
        
        # Update UI to show new sort state
        update_sort_ui()

    def map_filter(txt):
        m = {
            'جستجوی کلی': 'full',
            'بر اساس عنوان': 'title',
            'بر اساس موضوع': 'subject',
            'بر اساس تاریخ': 'date',
            'بر اساس نوع پرونده': 'case_type',
            'بر اساس شناسه بایگانی': 'id'
        }
        return m.get(txt, 'full')

    def do_search():
        ft = map_filter(combo.get())
        
        # Handle date range search
        if ft == 'date':
            date_from = entry_date_from.get().strip()
            date_to = entry_date_to.get().strip()
            
            if not date_from or not date_to:
                messagebox.showwarning('هشدار', 'لطفا هر دو تاریخ را وارد کنید', parent=top)
                return
            
            try:
                # Validate dates
                d_from = jdatetime.datetime.strptime(date_from, '%Y-%m-%d').date()
                d_to = jdatetime.datetime.strptime(date_to, '%Y-%m-%d').date()
                if d_from > d_to:
                    messagebox.showwarning('هشدار', 'تاریخ شروع باید قبل از تاریخ پایان باشد', parent=top)
                    return
            except Exception as e:
                messagebox.showerror('خطا', f'فرمت تاریخ نامعتبر: {e}', parent=top)
                return
            
            # Query database for cases in date range (include duration_from and duration_to)
            conn = get_connection()
            cur = conn.cursor()
            cur.execute('''SELECT id, title, subject, date, case_type, duration, status, contract_amount, duration_from, duration_to
                           FROM cases 
                           WHERE date >= ? AND date <= ? 
                           ORDER BY date DESC''', (date_from, date_to))
            rows = cur.fetchall()
            conn.close()
        else:
            # Regular search
            q = entry_q.get().strip()
            rows = search_cases(ft, q)
            # Fetch duration_from and duration_to for all results
            if rows:
                case_ids = [str(r[0]) for r in rows]
                conn = get_connection()
                cur = conn.cursor()
                cur.execute('SELECT id, duration_from, duration_to FROM cases WHERE id IN ({})'.format(','.join(['?' for _ in case_ids])), case_ids)
                duration_data = {row[0]: (row[1], row[2]) for row in cur.fetchall()}
                conn.close()
                # Add duration_from and duration_to to each row
                rows = [r + (duration_data.get(r[0], ('', ''))[0], duration_data.get(r[0], ('', ''))[1]) for r in rows]
        
        for i in tree.get_children():
            tree.delete(i)
        for r in rows:
            # r[0-7] = id, title, subject, date, case_type, duration, status, contract_amount
            # r[8-9] = duration_from, duration_to (from our queries above)
            duration_text = calculate_duration_text(r[8] if len(r) > 8 else None, r[9] if len(r) > 9 else None)
            vals = [
                str(r[0]) if r[0] else '-----',           # id
                str(r[1]) if r[1] else '-----',           # title
                str(r[2]) if r[2] else '-----',           # subject
                str(r[3]) if r[3] else '-----',           # date
                str(r[4]) if r[4] else '-----',           # case_type
                duration_text,                              # duration (calculated)
                str(r[6]) if r[6] else '-----',           # status
                (f"{int(r[7]):,} ریال" if r[7] else '-----') if isinstance(r[7], (int, float)) or (isinstance(r[7], str) and r[7].replace(',', '').isdigit()) else str(r[7]) if r[7] else '-----'  # contract_amount with ریال
            ]
            tree.insert('', 'end', values=vals)
        
        # Update status label with result count
        lbl_status.configure(text=f'تعداد نتایج یافت شده: {len(rows)}', text_color='green' if rows else 'gray')
        
        # Enable export buttons if there are results
        btn_export_xlsx.configure(state="normal" if rows else "disabled")
        
        # Store results for export
        current_data['rows'] = rows
        
        # Reset sort state
        sort_state['column'] = None
        sort_state['reverse'] = False
        # Reset radio buttons and clear sort indicators
        update_sort_ui()

    def on_open_details(event):
        sel = tree.selection()
        if not sel:
            return
        item = tree.item(sel[0])
        case_id = item['values'][0] # The case_id is the first value in the tuple
        open_details_window(top, case_id)

    def on_selection_change(event):
        """Enable/disable delete button based on selection."""
        if tree.selection():
            btn_delete.configure(state="normal")
        else:
            btn_delete.configure(state="disabled")

    def delete_selected_records():
        """Delete selected record(s) - handles both single and multiple selection."""
        selected_items = tree.selection()
        if not selected_items:
            return

        count = len(selected_items)
        
        # For single selection, show simple confirmation
        if count == 1:
            item = tree.item(selected_items[0])
            case_id = item['values'][0]
            if messagebox.askyesno('تایید حذف', f'آیا از حذف پرونده با شناسه {case_id} مطمئن هستید؟', parent=top):
                try:
                    delete_case(case_id)
                    messagebox.showinfo('موفق', 'پرونده با موفقیت حذف شد.', parent=top)
                    do_search() # Refresh the list
                except Exception as e:
                    messagebox.showerror('خطا', f'خطا در هنگام حذف: {e}', parent=top)
        else:
            # For multiple selection, show group delete confirmation
            if messagebox.askyesno('تایید حذف گروهی', f'آیا از حذف {count} پرونده مطمئن هستید؟\nاین عملیات قابل بازگشت نیست.', parent=top):
                deleted_count = 0
                failed_count = 0
                
                for item_id in selected_items:
                    try:
                        item = tree.item(item_id)
                        case_id = item['values'][0]
                        delete_case(case_id)
                        deleted_count += 1
                    except Exception as e:
                        failed_count += 1
                        print(f"خطا در حذف پرونده {item['values'][0]}: {e}")
                
                # Show result message
                if failed_count == 0:
                    messagebox.showinfo('موفق', f'{deleted_count} پرونده با موفقیت حذف شد.', parent=top)
                else:
                    messagebox.showwarning('تکمیل با خطا', f'{deleted_count} پرونده حذف شد اما {failed_count} پرونده حذف نشد.', parent=top)
                
                do_search() # Refresh the list

    tree.bind('<Double-1>', on_open_details)
    tree.bind('<<TreeviewSelect>>', on_selection_change)
    tree.bind('<Control-a>', lambda e: select_all_items())

    def select_all_items():
        """Select all items in the treeview."""
        for item in tree.get_children():
            tree.selection_add(item)
        # Trigger selection change event
        on_selection_change(None)

    def apply_sort_controls():
        """Apply sorting based on the radio button selections."""
        col = sort_col_var.get()
        reverse = sort_dir_var.get() == 'desc'
        
        # Update sort_state
        sort_state['column'] = col
        sort_state['reverse'] = reverse
        
        items = [(tree.set(k, col), k) for k in tree.get_children('')]
        
        # Try to sort as numbers first, fallback to strings
        try:
            items.sort(key=lambda x: float(x[0]) if x[0] != '-----' else float('-inf'), reverse=reverse)
        except ValueError:
            items.sort(key=lambda x: x[0], reverse=reverse)
        
        # Rearrange items in sorted position
        for index, (val, k) in enumerate(items):
            tree.move(k, '', index)
        
        # Update column headers to show sort direction
        update_sort_ui()

    def export_to_xlsx():
        """Export current search results to an XLSX file."""
        if not current_data['rows']:
            messagebox.showwarning('هشدار', 'جدولی برای خروجی موجود نیست', parent=top)
            return

        file_path = filedialog.asksaveasfilename(
            parent=top,
            defaultextension='.xlsx',
            filetypes=[('Excel Files', '*.xlsx'), ('All Files', '*.*')],
            initialfile=f"نتایج_جستجو_{jdatetime.date.today().strftime('%Y%m%d')}.xlsx"
        )

        if not file_path:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "نتایج جستجو"
            sheet.sheet_view.rightToLeft = True

            # --- Styles ---
            bold_font = Font(name='Vazirmatn', bold=True)
            regular_font = Font(name='Vazirmatn')
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
            # readingOrder=2 forces RTL direction in the cell, which fixes display on Windows
            center_alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)

            # --- Fetch full data for XLSX export ---
            case_ids = [str(row[0]) for row in current_data['rows']]
            conn = get_connection()
            cur = conn.cursor()
            cur.execute('SELECT * FROM cases WHERE id IN ({})'.format(','.join(['?' for _ in case_ids])), case_ids)
            full_rows = cur.fetchall()
            column_names = [desc[0] for desc in cur.description]
            conn.close()

            # --- Define Headers and Columns for XLSX ---
            column_labels = {
                'id': 'شناسه بایگانی',
                'title': 'عنوان',
                'date': 'تاریخ',
                'duration': 'مدت',
                'duration_from': 'تاریخ شروع',
                'duration_to': 'تاریخ پایان',
                'mojer': 'موجر',
                'mostajjer': 'مستاجر',
                'karfarma': 'کارفرما',
                'piman': 'پیمانکار',
                'subject': 'موضوع',
                'contract_amount': 'مبلغ قرارداد',
                'bank_owner_name': 'نام صاحب حساب',
                'bank_account_number': 'شماره حساب',
                'bank_shaba_number': 'شماره شبا',
                'bank_card_number': 'شماره کارت',
                'bank_name': 'نام بانک',
                'bank_branch': 'شعبه بانک',
                'payment_id': 'شناسه پرداخت',
                'guarantee_amount': 'مبلغ تضمین',
                'guarantee_type': 'نوع ضمانت',
                'description': 'توضیحات',
                'case_type': 'نوع پرونده',
                'status': 'وضعیت',
            }
            filtered_columns = [col for col in column_names if col not in ('folder_path', 'parties', 'duration')]
            duration_index = column_names.index('duration') if 'duration' in column_names else -1

            # --- Headers ---
            persian_headers = [column_labels.get(col, col) for col in filtered_columns]
            if duration_index != -1:
                persian_headers.insert(duration_index, 'مدت')
            persian_headers.insert(0, 'ردیف') # Add Row column header
            
            sheet.append(persian_headers)

            # Style Header
            for col_num, header in enumerate(persian_headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = bold_font
                cell.border = thin_border
                cell.alignment = center_alignment

            # --- Data Rows ---
            for row_idx, row in enumerate(full_rows, start=2):
                duration_from_idx = column_names.index('duration_from') if 'duration_from' in column_names else -1
                duration_to_idx = column_names.index('duration_to') if 'duration_to' in column_names else -1
                row_list = list(row)
                calculated_duration = '-----'
                if duration_from_idx != -1 and duration_to_idx != -1:
                    calculated_duration = calculate_duration_text(row_list[duration_from_idx], row_list[duration_to_idx])

                filtered_row = build_filtered_row(row_list, column_names, calculated_duration, add_rial=True)
                
                # Add row number
                filtered_row.insert(0, row_idx - 1)

                # Reshape for display
                reshaped_row = [str(item) if item is not None else '' for item in filtered_row]
                sheet.append(reshaped_row)

                # Apply font, border, and alignment to data cells
                for col_idx, _ in enumerate(reshaped_row, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.font = regular_font
                    cell.border = thin_border
                    cell.alignment = center_alignment

            # --- Adjust Column Widths ---
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width

            workbook.save(file_path)
            
            messagebox.showinfo('موفق', f'گزارش با موفقیت صادر شد:\n{file_path}', parent=top)
        except Exception as e:
            messagebox.showerror('خطا', f'خطا در صادر کردن فایل XLSX: {str(e)}', parent=top)

    def build_filtered_row(row_list, column_names, calculated_duration, add_rial=False):
        """Helper function to build a row for CSV/XLSX export."""
        filtered_row = []
        for i, col in enumerate(column_names):
            if col not in ('folder_path', 'parties'):
                if col == 'duration':
                    filtered_row.append(calculated_duration)
                elif col == 'contract_amount' and add_rial:
                    value = row_list[i]
                    if value and (isinstance(value, (int, float)) or (isinstance(value, str) and str(value).replace(',', '').isdigit())):
                        try:
                            filtered_row.append(f"{int(value):,} ریال")
                        except (ValueError, TypeError):
                            filtered_row.append(value)
                    else:
                        filtered_row.append(value)
                else:
                    filtered_row.append(row_list[i])
        return filtered_row

    # Define these here to be accessible by export functions
    duration_from_idx = -1
    duration_to_idx = -1

    # Set button commands
    btn_export_xlsx.configure(command=export_to_xlsx)

    tree.bind('<Double-1>', on_open_details)
    tree.bind('<<TreeviewSelect>>', on_selection_change)
    tree.bind('<Control-a>', lambda e: select_all_items())

    def select_all_items():
        """Select all items in the treeview."""
        for item in tree.get_children():
            tree.selection_add(item)
        # Trigger selection change event
        on_selection_change(None)

    # Load all records on startup
    do_search()
